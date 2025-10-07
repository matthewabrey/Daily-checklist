#!/usr/bin/env python3
import asyncio
import os
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

async def debug_checklist_update():
    # MongoDB setup
    MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
    DB_NAME = os.environ.get("DB_NAME", "test_database")
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    
    # Load the Excel file
    file_path = "/tmp/AssetList_latest.xlsx"
    workbook = load_workbook(file_path)
    
    print("📊 Excel File Analysis:")
    print(f"Sheet names: {workbook.sheetnames}")
    
    # Get asset check types
    sheet = workbook.active
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
    print(f"Headers: {headers}")
    
    check_type_col = None
    for i, header in enumerate(headers):
        if 'check type' in header or 'checktype' in header or header == 'check type':
            check_type_col = i
            break
    
    if check_type_col is None:
        print("❌ Could not find check type column")
        return
    
    # Get unique check types
    unique_check_types = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and len(row) > check_type_col and row[check_type_col]:
            check_type = str(row[check_type_col]).strip()
            unique_check_types.add(check_type)
    
    print(f"Unique check types found: {list(unique_check_types)}")
    
    # Process each sheet
    print("\n🔍 Processing Sheets:")
    for sheet_name in workbook.sheetnames:
        if sheet_name == workbook.sheetnames[0]:
            print(f"  Skipping main sheet: {sheet_name}")
            continue
            
        sheet = workbook[sheet_name]
        print(f"\n  📋 Processing: {sheet_name}")
        
        # Try to match with check types
        sheet_name_clean = sheet_name.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '')
        matching_check_type = None
        
        for check_type in unique_check_types:
            check_type_clean = check_type.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '')
            if check_type_clean in sheet_name_clean or sheet_name_clean in check_type_clean:
                matching_check_type = check_type
                break
        
        if not matching_check_type:
            matching_check_type = sheet_name
            
        print(f"    Sheet name clean: '{sheet_name_clean}'")
        print(f"    Matched to: '{matching_check_type}'")
        
        # Extract items
        items = []
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_num == 1:  # Skip header
                continue
                
            if row and row[0]:
                item_text = str(row[0]).strip()
                if (item_text and 
                    item_text.lower() not in ['item', 'check', 'description', 'checklist', 'safety'] and
                    len(item_text) > 3):
                    items.append(item_text)
        
        print(f"    Items found: {len(items)}")
        if items:
            print(f"    First few items: {items[:3]}")

if __name__ == "__main__":
    asyncio.run(debug_checklist_update())