#!/usr/bin/env python3
from openpyxl import load_workbook

def debug_updated_file():
    file_path = "/tmp/AssetList_updated.xlsx"
    workbook = load_workbook(file_path)
    
    print("📊 Sheet Analysis:")
    print(f"All sheets: {workbook.sheetnames}")
    
    # Check the first sheet (main asset sheet)
    first_sheet = workbook[workbook.sheetnames[0]]
    print(f"\nFirst sheet name: {workbook.sheetnames[0]}")
    
    # Get headers
    headers = []
    for cell in first_sheet[1]:
        headers.append(str(cell.value).strip() if cell.value else '')
    
    print(f"Headers in first sheet: {headers}")
    
    # Check first few rows of data
    print(f"\nFirst 5 data rows:")
    for i, row in enumerate(first_sheet.iter_rows(min_row=2, max_row=6, values_only=True)):
        print(f"Row {i+2}: {row}")
    
    # Check the Forklift JCB sheet specifically
    if 'Forklift JCB' in workbook.sheetnames:
        forklift_sheet = workbook['Forklift JCB']
        print(f"\n🚀 Forklift JCB Sheet Analysis:")
        
        # Get all items from first column
        items = []
        for row in forklift_sheet.iter_rows(values_only=True):
            if row and row[0]:
                item_text = str(row[0]).strip()
                items.append(item_text)
        
        print(f"Total items: {len(items)}")
        print("Items containing 'working' or ending with '1':")
        
        for item in items:
            if 'working' in item.lower() or item.endswith('1'):
                print(f"  - {item}")

if __name__ == "__main__":
    debug_updated_file()