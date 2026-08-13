from fastapi import FastAPI, HTTPException, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import os
import io
from motor.motor_asyncio import AsyncIOMotorClient
import uuid
from bson import ObjectId
from dotenv import load_dotenv
from sharepoint_integration import sharepoint_integration
from sharepoint_auto_sync import sharepoint_auto_sync
from cached_stats import get_cached_stats, invalidate_cache
from fieldplan_sync import download_fieldplan, FIELDPLAN_PATH, download_fieldmap, FIELDMAP_PATH
import qrcode
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import asyncio
import logging
import httpx

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

app = FastAPI(title="Machine Checklist API")

# Scheduler for automatic SharePoint sync
scheduler = AsyncIOScheduler(timezone="Europe/London")

# CORS setup
CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "*").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB setup with connection pooling and timeouts
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")

# Configure MongoDB client with performance settings
client = AsyncIOMotorClient(
    MONGO_URL,
    maxPoolSize=10,  # Connection pool size
    minPoolSize=1,
    maxIdleTimeMS=30000,  # Close idle connections after 30s
    serverSelectionTimeoutMS=5000,  # Fail fast if can't connect
    connectTimeoutMS=10000,  # Connection timeout
    socketTimeoutMS=30000,  # Socket timeout for queries
)
db = client[DB_NAME]

# Collections
# db.checklists - checklist records
# db.assets - machine/asset data
# db.staff - staff data
# db.repair_status - tracks acknowledged/completed status of repairs (NEW)
# db.sync_logs - SharePoint sync history

# Scheduled SharePoint sync function
async def scheduled_sharepoint_sync():
    """Scheduled task to sync staff list and assets from SharePoint at 9am daily"""
    logger.info("Starting scheduled SharePoint sync (staff + assets)...")
    try:
        result = await sharepoint_auto_sync.sync_all(db)
        # Log the sync result
        await db.sync_logs.insert_one({
            'type': 'scheduled',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'success': result.get('success', False),
            'message': f"Staff: {result.get('staff', {}).get('message', 'N/A')}, Assets: {result.get('assets', {}).get('message', 'N/A')}",
            'staff_count': result.get('staff', {}).get('count', 0),
            'assets_count': result.get('assets', {}).get('assets_count', 0),
            'templates_count': result.get('assets', {}).get('templates_count', 0)
        })
        logger.info(f"Scheduled sync completed: Staff={result.get('staff', {}).get('success')}, Assets={result.get('assets', {}).get('success')}")
    except Exception as e:
        logger.error(f"Scheduled sync error: {str(e)}")
        await db.sync_logs.insert_one({
            'type': 'scheduled',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'success': False,
            'message': str(e),
            'staff_count': 0,
            'assets_count': 0
        })

# Scheduled FieldPlan (cropping map) download
async def scheduled_fieldplan_sync():
    """Scheduled task to re-download the external cropping FieldPlan daily"""
    try:
        size = await download_fieldplan()
        logger.info(f"Scheduled FieldPlan sync completed ({size} chars)")
    except Exception as e:
        logger.error(f"Scheduled FieldPlan sync error: {str(e)}")
    try:
        size = await download_fieldmap()
        logger.info(f"Scheduled FieldMap sync completed ({size} chars)")
    except Exception as e:
        logger.error(f"Scheduled FieldMap sync error: {str(e)}")

# Setup scheduler on startup
@app.on_event("startup")
async def startup_event():
    """Start the scheduler when the app starts"""
    # Schedule daily sync at 9:00 AM UK time
    scheduler.add_job(
        scheduled_sharepoint_sync,
        CronTrigger(hour=9, minute=0, timezone="Europe/London"),
        id="daily_staff_sync",
        name="Daily SharePoint Staff Sync",
        replace_existing=True
    )
    # Schedule daily FieldPlan download at 5:00 AM UK time
    scheduler.add_job(
        scheduled_fieldplan_sync,
        CronTrigger(hour=5, minute=0, timezone="Europe/London"),
        id="daily_fieldplan_sync",
        name="Daily FieldPlan Map Download",
        replace_existing=True
    )
    scheduler.start()
    logger.info("Scheduler started - Daily staff sync scheduled for 9:00 AM UK time")
    # Download the FieldPlan/FieldMap immediately if we don't have a copy yet
    if not os.path.exists(FIELDPLAN_PATH) or not os.path.exists(FIELDMAP_PATH):
        asyncio.create_task(scheduled_fieldplan_sync())

@app.on_event("shutdown")
async def shutdown_event():
    """Stop the scheduler when the app shuts down"""
    scheduler.shutdown()
    logger.info("Scheduler stopped")

@app.get("/api/fieldplan")
async def get_fieldplan():
    """Serve the self-hosted copy of the external cropping FieldPlan map"""
    if not os.path.exists(FIELDPLAN_PATH):
        try:
            await download_fieldplan()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Failed to fetch field plan: {str(e)}")
    return FileResponse(FIELDPLAN_PATH, media_type="text/html")

@app.get("/api/fieldmap")
async def get_fieldmap():
    """Serve the self-hosted copy of the external FieldMap (map with filters)"""
    if not os.path.exists(FIELDMAP_PATH):
        try:
            await download_fieldmap()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Failed to fetch field map: {str(e)}")
    return FileResponse(FIELDMAP_PATH, media_type="text/html")

@app.post("/api/fieldplan/refresh")
async def refresh_fieldplan():
    """Manually re-download the latest FieldPlan and FieldMap from the external site"""
    try:
        size = await download_fieldplan()
        map_size = await download_fieldmap()
        return {"success": True, "size": size, "map_size": map_size}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to refresh field plan: {str(e)}")

# Pydantic models
class Asset(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    check_type: str
    name: str  # Name of Implement
    make: str
    qr_printed: bool = False  # Whether QR code has been printed for this asset
    qr_printed_at: Optional[str] = None  # ISO timestamp when QR was printed
    
class Staff(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_number: Optional[str] = None
    name: str
    active: bool = True
    workshop_control: Optional[str] = None  # "yes" or "no" from Workshop Control column
    admin_control: Optional[str] = None  # "yes" or "no" from Admin Control column
    manager_control: Optional[str] = None  # "yes" or "no" from Manager Control column
    
class ChecklistItem(BaseModel):
    item: str
    status: str = "unchecked"  # "unchecked", "satisfactory", "unsatisfactory", "n/a"
    notes: Optional[str] = None
    photos: Optional[List[dict]] = []
    compulsory: bool = False  # If True, item cannot be marked unsatisfactory when signing off
    
class ChecklistTemplateItem(BaseModel):
    item: str
    compulsory: bool = False

class ChecklistTemplate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    check_type: str  # "daily_check", "grader_startup", "workshop_service"
    items: List[ChecklistTemplateItem]  # Now includes compulsory flag per item
    
class Checklist(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_number: str
    staff_name: str
    machine_make: str
    machine_model: str
    check_type: str  # "daily_check", "grader_startup", "workshop_service", or "fuel_mileage"
    checklist_items: List[ChecklistItem] = []
    workshop_notes: Optional[str] = None
    workshop_photos: Optional[List[dict]] = []
    # Fuel and Mileage fields
    fuel_mileage: Optional[str] = None
    fuel_added: Optional[str] = None
    adblue_added: Optional[str] = None
    fuel_notes: Optional[str] = None
    completed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "completed"
    
class ChecklistResponse(BaseModel):
    id: str
    employee_number: Optional[str] = None
    staff_name: str
    machine_make: str
    machine_model: str
    check_type: str
    checklist_items: List[ChecklistItem]
    workshop_notes: Optional[str] = None
    workshop_photos: Optional[List[dict]] = []
    # Fuel and Mileage fields
    fuel_mileage: Optional[str] = None
    fuel_added: Optional[str] = None
    adblue_added: Optional[str] = None
    fuel_notes: Optional[str] = None
    completed_at: datetime
    status: str

# Work Progress Tracking Models
class WorkEntry(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_id: str
    hectares_completed: float
    date_completed: str  # ISO date string
    entered_by: str  # Employee name
    entered_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Job(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # e.g., "Carrot Drilling"
    total_area: float  # Total hectares
    target_date: Optional[str] = None  # Target completion date (YYYY-MM-DD)
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    status: str = "active"  # "active" or "complete"

class JobCreate(BaseModel):
    name: str
    total_area: float
    target_date: Optional[str] = None  # Target completion date (YYYY-MM-DD)

class WorkEntryCreate(BaseModel):
    hectares_completed: float
    date_completed: Optional[str] = None  # If not provided, use today
    entered_by: str

# Near Miss and Suggestion Models
class NearMiss(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str
    location: Optional[str] = None
    photos: List[str] = []  # Base64 encoded photos
    is_anonymous: bool = False
    submitted_by: Optional[str] = None  # Name if not anonymous
    employee_number: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    acknowledged: bool = False
    acknowledged_at: Optional[str] = None
    acknowledged_by: Optional[str] = None
    comments: List[dict] = []
    # Investigation fields
    severity: Optional[str] = None  # red, orange, green
    action_required: Optional[str] = None
    progress: Optional[str] = None  # not_started, in_progress, completed
    investigation_notes: Optional[str] = None
    no_swp_or_not_covered: bool = False  # No Safe Working Procedure or it doesn't cover this
    swp_training_not_received: bool = False  # Training on SWP not received by person
    trained_but_not_following: bool = False  # Trained but individual not following SWP
    investigated_by: Optional[str] = None
    investigated_at: Optional[str] = None

class NearMissCreate(BaseModel):
    description: str
    location: Optional[str] = None
    photos: List[str] = []
    is_anonymous: bool = False
    submitted_by: Optional[str] = None

class Suggestion(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    category: Optional[str] = None  # Financial, Well Being, Health and Safety
    location: Optional[str] = None  # Farm, Field, Storage, Grading
    photos: List[str] = []  # Base64 encoded photos
    is_anonymous: bool = False
    submitted_by: Optional[str] = None
    employee_number: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    status: str = "new"  # new, reviewed, implemented, declined
    reviewed_at: Optional[str] = None
    reviewed_by: Optional[str] = None
    review_notes: Optional[str] = None

class SuggestionCreate(BaseModel):
    title: str
    description: str
    category: Optional[str] = None
    location: Optional[str] = None
    photos: List[str] = []
    is_anonymous: bool = False
    submitted_by: Optional[str] = None

# Accident Models - Matching official accident record book
class Accident(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    report_number: Optional[str] = None
    
    # Section 1: About the person who had the accident
    injured_name: str
    injured_address: Optional[str] = None
    injured_postcode: Optional[str] = None
    injured_occupation: Optional[str] = None
    
    # Section 2: About you, the person filling in this record
    reporter_name: str
    reporter_address: Optional[str] = None
    reporter_postcode: Optional[str] = None
    reporter_occupation: Optional[str] = None
    
    # Section 3: About the accident
    accident_date: str  # Date of accident
    accident_time: str  # Time of accident
    accident_location: str  # Where it happened (room or place)
    accident_description: str  # How the accident happened / cause
    injury_details: Optional[str] = None  # What injury was suffered
    
    # Signature section
    signature_date: Optional[str] = None
    
    # Section 4: For the employee only - consent
    employee_consent: bool = False
    employee_signature_date: Optional[str] = None
    
    # Section 5: For the employer only - RIDDOR
    riddor_reportable: bool = False
    riddor_how_reported: Optional[str] = None
    riddor_date_reported: Optional[str] = None
    employer_signature: Optional[str] = None
    
    # Additional fields
    photos: List[str] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    status: str = "new"  # new, investigating, closed
    comments: List[dict] = []
    investigation_notes: Optional[str] = None
    investigated_by: Optional[str] = None
    investigated_at: Optional[str] = None

class AccidentCreate(BaseModel):
    # Section 1
    injured_name: str
    injured_address: Optional[str] = None
    injured_postcode: Optional[str] = None
    injured_occupation: Optional[str] = None
    
    # Section 2
    reporter_name: str
    reporter_address: Optional[str] = None
    reporter_postcode: Optional[str] = None
    reporter_occupation: Optional[str] = None
    
    # Section 3
    accident_date: str
    accident_time: str
    accident_location: str
    accident_description: str
    injury_details: Optional[str] = None
    
    # Section 4
    employee_consent: bool = False
    
    # Photos
    photos: List[str] = []

# Whistleblowing Models
class Whistleblow(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    category: Optional[str] = None  # Financial, Health and Safety, Misconduct, Other
    location: Optional[str] = None
    is_anonymous: bool = True  # Default anonymous for whistleblowing
    submitted_by: Optional[str] = None
    employee_number: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    status: str = "new"  # new, investigating, resolved, dismissed
    investigated_at: Optional[str] = None
    investigated_by: Optional[str] = None
    investigation_notes: Optional[str] = None
    comments: List[dict] = []

class WhistleblowCreate(BaseModel):
    title: str
    description: str
    category: Optional[str] = None
    location: Optional[str] = None
    is_anonymous: bool = True
    submitted_by: Optional[str] = None

# Training Record Models
class TraineeSignature(BaseModel):
    employee_id: Optional[str] = None
    employee_name: str
    is_agency: bool = False
    signed: bool = False
    signed_at: Optional[str] = None
    signature_data: Optional[str] = None  # Base64 signature image

class TrainingRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    swp_number: str  # Safe Working Practice number
    swp_version: str
    department: str
    training_date: str
    notes: Optional[str] = None
    trainer_name: str
    trainer_employee_number: Optional[str] = None
    trainees: List[TraineeSignature] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    status: str = "pending_signatures"  # pending_signatures, completed
    added_to_sage_hr: bool = False
    added_to_sage_hr_at: Optional[str] = None
    added_to_sage_hr_by: Optional[str] = None

class TrainingRecordCreate(BaseModel):
    swp_number: str
    swp_version: str
    department: str
    training_date: str
    notes: Optional[str] = None
    trainer_name: str
    trainer_employee_number: Optional[str] = None
    trainees: List[dict] = []  # List of {employee_id, employee_name, is_agency}

# Initialize data
async def initialize_data():
    # Check if data already exists
    asset_count = await db.assets.count_documents({})
    staff_count = await db.staff.count_documents({})
    
    # Skip asset initialization - assets should be uploaded via Admin Panel
    # using AssetList.xlsx with proper format (Check Type | Name | Make)
    if asset_count == 0:
        print("No assets found. Please upload AssetList.xlsx via Admin Panel.")
    
    if staff_count == 0:
        # Staff data from Excel
        staff_names = [
            "Abbie Nixon", "Adrian-Stefan Iovu", "Alan Day", "Andrew Rose", "Andy Browning",
            "Angele Ganitauskiene", "Angele Samuliene", "Armen Saakov", "Audrius Ambramavicius",
            "Biser Aleksiev", "Biser Borisov", "Bonnie Oakley", "Caitlin Barnes",
            "Christopher Marsh", "Clive Bowman", "Cristinel Susma", "Dimitar Boev",
            "Dumitru Verdes", "Edis Daud", "Florin Iovu", "Gary Harrowing",
            "Gheorghe Caraman", "Gina Caraman", "Ginka Koleva", "Hayden Bennett",
            "Hazel Cassinelli", "Hristo Samardzhiev", "Ion Dascal", "Iurii Cujba",
            "Jake Murfitt", "James Butler", "James Rogerson", "Jerry Langridge",
            "Jon Pearson", "Julian Ingleson", "Kasim Yusein", "Kieran Button",
            "Kieran Rushbrook", "Kostadin Stoyanov", "Lilyana Babakova", "Lina Barkauskyte",
            "Maclafan Mugova", "Marcus Patience", "Mark Belton", "Mike Cameron",
            "Milka Nankova", "Nilyay Yovu-Saami", "Nurten Yusein", "Paul Churchyard",
            "Paul Dye", "Rafal Trela", "Razvan Laszlo", "Rhys Le-Gallez",
            "Richard Stennett", "Robert Northwood", "Samuel Watkins", "Shevked Halibryam",
            "Simon Denton", "Stephen Tortice", "Stoyan Stoyanov", "Stuart Hatch",
            "Tamas Kishajdu", "Tanta Laszlo", "Terry Davidson", "Tomas Urbutis",
            "Tommy Kefford", "Victoria Dascal", "Violeta Stoyanova", "Zander Britton"
        ]
        
        for staff_name in staff_names:
            staff = Staff(name=staff_name)
            staff_dict = staff.dict()
            await db.staff.insert_one(staff_dict)
            
        # Add admin employee 
        admin_staff = Staff(employee_number="4444", name="Admin User", admin_control="yes", manager_control="yes")
        admin_dict = admin_staff.dict()
        await db.staff.insert_one(admin_dict)

async def initialize_workplan_data():
    """Seed default jobs and colour categories for the Daily Workplan feature."""
    if await db.workplan_jobs.count_documents({}) == 0:
        default_jobs = [
            "AD", "Accommodation", "Band Spraying", "Bed Tilling", "Bed Cracker",
            "Bed Mixing", "Bed Chopping", "Box Sorting", "Due Back IN", "Bowser",
            "Carting Crop", "Carting Manure", "Chitting Seed", "Chickens", "Course",
            "Dam Diking", "De Bagging", "Deep Cultivation", "Destoning", "Digging",
            "Drain Jetting", "Drilling", "Estate Work", "Flailing", "Fencing", "360 Work",
            "Fertiliser Spreading", "Night Destoning", "Fleecing", "Forklift", "Finished",
            "Grading", "Harvest", "Hedge Cutting", "Headland Cultivation", "Hoeing",
            "Holiday", "MOTHERS DAY", "Irrigation Moving", "Irrigation Overground",
            "Irrigation Service", "Irrigation Underground", "Loadall", "Keeble",
            "Loading - JCB", "Off", "Muck Spreading", "Office Work", "Rolling", "PARTY",
            "Planting", "Ploughing", "Poly Laying", "Poly Lifting", "Packing", "On Call",
            "re-ridging", "Rotavating", "Ridging", "Sample Digging", "Seed",
            "Service Cultivation", "Service Harvest", "Service Other",
            "Service Planting/Drill", "Shallow Cultivation", "Shooting", "Sickness",
            "Spraying", "Storage", "Stone Burying", "Tractor Shed", "Tractor Only Broken",
            "Training", "TWB", "Wind rowing", "Wheel Change", "Weeding", "Topping",
            "Tractor Only Spare", "Fert bowser", "QC", "Irrigation",
            "Wet Day Jobs", "Webb Repair", "Shelfs To Sort In Workshop",
            "Clean Work Vans", "Paint Rest Room / Toitlets", "Sort Camp Pad Out",
            "Thorpe Farm Tidy", "Poly / Fleece Sort out", "Trees over larkshall fence"
        ]
        for i, name in enumerate(default_jobs):
            await db.workplan_jobs.insert_one({"id": str(uuid.uuid4()), "name": name, "order": i})

    if await db.workplan_colors.count_documents({}) == 0:
        default_colors = [
            ("Onions", "#16a34a"),
            ("Carrots", "#f97316"),
            ("Potatoes", "#a16207"),
            ("Larkshall", "#0ea5e9"),
            ("Snetterton", "#a855f7"),
            ("Off / Holiday", "#ef4444"),
            ("Servicing", "#eab308"),
        ]
        for i, (name, color) in enumerate(default_colors):
            await db.workplan_colors.insert_one({"id": str(uuid.uuid4()), "name": name, "color": color, "order": i})

@app.on_event("startup")
async def startup_event():
    await initialize_data()
    await initialize_workplan_data()
    await migrate_existing_checklists()
    await ensure_indexes()

async def ensure_indexes():
    """Ensure all required indexes exist for performance"""
    try:
        print("Ensuring database indexes...")
        
        # Checklists indexes - optimized for common queries
        await db.checklists.create_index([("completed_at", -1)])
        await db.checklists.create_index([("check_type", 1)])
        await db.checklists.create_index([("check_type", 1), ("completed_at", -1)])
        await db.checklists.create_index([("machine_make", 1)])
        await db.checklists.create_index([("machine_make", 1), ("completed_at", -1)])  # For by-machine queries
        await db.checklists.create_index([("machine_make", 1), ("machine_model", 1), ("completed_at", -1)])  # Compound index
        await db.checklists.create_index([("employee_number", 1)])
        await db.checklists.create_index([("id", 1)])
        await db.checklists.create_index([("checklist_items.status", 1)])
        
        # Assets indexes
        await db.assets.create_index([("make", 1)])
        await db.assets.create_index([("make", 1), ("name", 1)])
        
        # Staff indexes
        await db.staff.create_index([("employee_number", 1)])
        await db.staff.create_index([("active", 1)])
        
        # Repair status indexes
        await db.repair_status.create_index([("repair_id", 1)])
        await db.repair_status.create_index([("acknowledged", 1)])
        await db.repair_status.create_index([("completed", 1)])
        
        print("Database indexes ensured successfully")
    except Exception as e:
        print(f"Warning: Could not create some indexes: {e}")

async def migrate_existing_checklists():
    """Add check_type field to existing checklists that don't have it"""
    try:
        # Find checklists without check_type field
        checklists_to_update = await db.checklists.find({"check_type": {"$exists": False}}).to_list(length=5000)
        
        if checklists_to_update:
            print(f"Migrating {len(checklists_to_update)} existing checklists...")
            
            # Update each checklist to add check_type as "daily_check" (assuming old records were daily checks)
            for checklist in checklists_to_update:
                # Convert old "checked" boolean to new "status" format
                if checklist.get('checklist_items'):
                    for item in checklist['checklist_items']:
                        if 'checked' in item and 'status' not in item:
                            item['status'] = 'satisfactory' if item['checked'] else 'unchecked'
                            item.pop('checked', None)  # Remove old field
                
                await db.checklists.update_one(
                    {"_id": checklist["_id"]}, 
                    {"$set": {
                        "check_type": "daily_check",
                        "workshop_notes": None,
                        "checklist_items": checklist.get('checklist_items', [])
                    }}
                )
            print(f"Successfully migrated {len(checklists_to_update)} checklists")
    except Exception as e:
        print(f"Migration error: {e}")

async def cleanup_duplicate_staff():
    """Remove duplicate staff entries, keeping the one with most permissions"""
    try:
        # Find all employee numbers with duplicates
        pipeline = [
            {"$group": {"_id": "$employee_number", "count": {"$sum": 1}, "ids": {"$push": "$id"}}},
            {"$match": {"count": {"$gt": 1}}}
        ]
        duplicates = await db.staff.aggregate(pipeline).to_list(length=100)
        
        for dup in duplicates:
            emp_num = dup["_id"]
            if not emp_num:
                continue
                
            # Get all records for this employee
            records = await db.staff.find({"employee_number": emp_num}).to_list(length=100)
            
            # Keep the one with admin_control='yes' or workshop_control='yes', or the first one
            best_record = None
            for r in records:
                if r.get("admin_control") == "yes" or r.get("workshop_control") == "yes":
                    best_record = r
                    break
            if not best_record:
                best_record = records[0]
            
            # Delete all except the best one
            for r in records:
                if r["_id"] != best_record["_id"]:
                    await db.staff.delete_one({"_id": r["_id"]})
            
            print(f"Cleaned up duplicates for employee {emp_num}")
    except Exception as e:
        print(f"Duplicate cleanup error: {e}")

# API Routes
@app.get("/api/health")
async def health_check():
    """Health check with database connectivity test"""
    try:
        # Quick database ping
        await db.command("ping")
        return {
            "status": "healthy",
            "database": "connected",
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

class EmployeeLoginRequest(BaseModel):
    employee_number: str

@app.post("/api/auth/employee-login")
async def employee_login(request: EmployeeLoginRequest):
    """Authenticate employee by number"""
    try:
        # Find employee by number
        employee = await db.staff.find_one({
            "employee_number": request.employee_number,
            "active": True
        }, {"_id": 0})
        
        print(f"[DEBUG] employee_login: Looking up {request.employee_number}")
        print(f"[DEBUG] employee_login: Found employee: {employee}")
        
        if employee:
            result = {
                "success": True,
                "employee": {
                    "employee_number": employee["employee_number"],
                    "name": employee["name"],
                    "workshop_control": employee.get("workshop_control", None),
                    "admin_control": employee.get("admin_control", None),
                    "manager_control": employee.get("manager_control", None)
                }
            }
            print(f"[DEBUG] employee_login: Returning: {result}")
            return result
        else:
            raise HTTPException(status_code=401, detail="Invalid employee number or account inactive")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Login failed: {str(e)}")

@app.get("/api/auth/validate/{employee_number}")
async def validate_employee(employee_number: str):
    """Validate if employee number is active"""
    try:
        employee = await db.staff.find_one({
            "employee_number": employee_number,
            "active": True
        }, {"_id": 0})
        
        if employee:
            return {
                "valid": True, 
                "name": employee["name"],
                "workshop_control": employee.get("workshop_control", None),
                "admin_control": employee.get("admin_control", None),
                "manager_control": employee.get("manager_control", None)
            }
        else:
            return {"valid": False}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Validation failed: {str(e)}")

@app.get("/api/debug/check-admin")
async def debug_check_admin():
    """Debug endpoint to check admin account workshop_control - REMOVE IN PRODUCTION"""
    try:
        admin = await db.staff.find_one({"employee_number": "4444"}, {"_id": 0})
        db_name = os.environ.get("DB_NAME", "not_set")
        return {
            "db_name": db_name,
            "admin_found": admin is not None,
            "admin_data": admin if admin else None,
            "code_version": "2024-11-11-v3-navigation-buttons"
        }
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/admin/deactivate-employee/{employee_number}")
async def deactivate_employee(employee_number: str):
    """Deactivate employee (block access)"""
    try:
        result = await db.staff.update_one(
            {"employee_number": employee_number},
            {"$set": {"active": False}}
        )
        
        if result.modified_count > 0:
            return {"message": f"Employee {employee_number} deactivated successfully"}
        else:
            raise HTTPException(status_code=404, detail="Employee not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to deactivate employee: {str(e)}")

@app.post("/api/admin/activate-employee/{employee_number}")
async def activate_employee(employee_number: str):
    """Reactivate employee"""
    try:
        result = await db.staff.update_one(
            {"employee_number": employee_number},
            {"$set": {"active": True}}
        )
        
        if result.modified_count > 0:
            return {"message": f"Employee {employee_number} activated successfully"}
        else:
            raise HTTPException(status_code=404, detail="Employee not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to activate employee: {str(e)}")

@app.post("/api/admin/grant-admin/{employee_number}")
async def grant_admin_access(employee_number: str):
    """Grant admin and workshop control to an employee"""
    try:
        # Update ALL documents with this employee number
        result = await db.staff.update_many(
            {"employee_number": employee_number},
            {"$set": {"admin_control": "yes", "workshop_control": "yes", "manager_control": "yes"}}
        )
        
        if result.modified_count > 0:
            return {"message": f"Admin access granted to {employee_number}", "modified": result.modified_count}
        else:
            raise HTTPException(status_code=404, detail="Employee not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to grant admin access: {str(e)}")

@app.get("/api/admin/employee-activity")
async def get_employee_activity():
    """Get employee usage statistics"""
    try:
        # Get recent checklists with employee info (last 90 days for performance)
        ninety_days_ago = (datetime.now(timezone.utc) - timedelta(days=90)).isoformat()
        checklists = await db.checklists.find(
            {"completed_at": {"$gte": ninety_days_ago}}, 
            {"_id": 0, "employee_number": 1, "staff_name": 1, "completed_at": 1}
        ).to_list(length=10000)
        
        # Count activity by employee
        activity = {}
        for checklist in checklists:
            emp_num = checklist.get('employee_number', 'Unknown')
            if emp_num not in activity:
                activity[emp_num] = {
                    "employee_number": emp_num,
                    "staff_name": checklist.get('staff_name', 'Unknown'),
                    "total_checks": 0,
                    "last_activity": None
                }
            
            activity[emp_num]["total_checks"] += 1
            
            # Update last activity
            completed_at = checklist.get('completed_at')
            if completed_at and (not activity[emp_num]["last_activity"] or completed_at > activity[emp_num]["last_activity"]):
                activity[emp_num]["last_activity"] = completed_at
        
        return list(activity.values())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get employee activity: {str(e)}")

@app.get("/api/staff", response_model=List[Staff])
async def get_staff():
    staff_list = await db.staff.find({}, {"_id": 0}).to_list(length=1000)  # Max 1000 staff
    return staff_list

@app.get("/api/assets/makes", response_model=List[str])
async def get_makes():
    makes = await db.assets.distinct("make")
    return sorted(makes)

@app.get("/api/assets/names/{make}", response_model=List[str])
async def get_names_by_make(make: str):
    names = await db.assets.distinct("name", {"make": make})
    return sorted(names)

@app.get("/api/assets/checktype/{make}/{name:path}")
async def get_checktype_by_make_and_name(make: str, name: str):
    asset = await db.assets.find_one({"make": make, "name": name}, {"_id": 0})
    if asset:
        check_type = asset["check_type"]
        # Handle nested check_type objects (from old data format)
        if isinstance(check_type, dict) and "check_type" in check_type:
            check_type = check_type["check_type"]
        return {"check_type": check_type}
    else:
        raise HTTPException(status_code=404, detail="Asset not found")

@app.get("/api/assets", response_model=List[Asset])
async def get_all_assets():
    assets = await db.assets.find({}, {"_id": 0}).to_list(length=1000)  # Max 1000 assets
    return assets

@app.get("/api/assets/qr-labels")
async def get_all_qr_labels():
    """Get list of all assets with QR code URLs and print status for printing page"""
    assets = await db.assets.find({}, {"_id": 0}).to_list(length=10000)
    
    # Add QR code URL to each asset and ensure qr_printed field exists
    for asset in assets:
        asset["qr_url"] = f"/api/assets/qr/{asset.get('make', '')}/{asset.get('name', '')}"
        # Ensure qr_printed field exists (for backward compatibility)
        if 'qr_printed' not in asset:
            asset['qr_printed'] = False
        if 'qr_printed_at' not in asset:
            asset['qr_printed_at'] = None
    
    return assets

@app.get("/api/assets/qr/{make}/{name}")
async def get_asset_qr_code(make: str, name: str):
    """Generate QR code for a machine"""
    # Create QR code data
    qr_data = f"MACHINE:{make}:{name}"
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    # Create image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to bytes
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    
    return StreamingResponse(img_bytes, media_type="image/png")

@app.get("/api/assets/{asset_id}")
async def get_asset_by_id(asset_id: str):
    """Get a single asset by ID"""
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset

@app.post("/api/assets/mark-qr-printed")
async def mark_assets_qr_printed(asset_ids: List[str]):
    """Mark multiple assets as having their QR codes printed"""
    timestamp = datetime.now(timezone.utc).isoformat()
    
    result = await db.assets.update_many(
        {"id": {"$in": asset_ids}},
        {"$set": {"qr_printed": True, "qr_printed_at": timestamp}}
    )
    
    return {
        "success": True,
        "modified_count": result.modified_count,
        "timestamp": timestamp
    }

@app.post("/api/assets/reset-qr-status")
async def reset_asset_qr_status(asset_ids: List[str]):
    """Reset QR printed status for specified assets (useful if labels need reprinting)"""
    result = await db.assets.update_many(
        {"id": {"$in": asset_ids}},
        {"$set": {"qr_printed": False, "qr_printed_at": None}}
    )
    
    return {
        "success": True,
        "modified_count": result.modified_count
    }

@app.post("/api/checklists", response_model=ChecklistResponse)
async def create_checklist(checklist: Checklist):
    # Validate compulsory items - if any compulsory item is marked unsatisfactory, reject the checklist
    if checklist.checklist_items:
        failed_compulsory_items = []
        for item in checklist.checklist_items:
            if item.compulsory and item.status == 'unsatisfactory':
                failed_compulsory_items.append(item.item)
        
        if failed_compulsory_items:
            item_list = ", ".join(failed_compulsory_items[:3])  # Show first 3 items
            if len(failed_compulsory_items) > 3:
                item_list += f" and {len(failed_compulsory_items) - 3} more"
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot sign off: Compulsory check(s) failed: {item_list}. Please resolve these issues before completing the checklist."
            )
    
    checklist_dict = checklist.dict()
    checklist_dict['completed_at'] = checklist_dict['completed_at'].isoformat()
    await db.checklists.insert_one(checklist_dict)
    
    # Invalidate dashboard cache so new machine additions show immediately
    await invalidate_cache()
    
    return ChecklistResponse(**checklist.dict())

@app.get("/api/dashboard/stats")
async def get_dashboard_stats():
    """ULTRA-FAST cached dashboard stats - returns in <50ms"""
    return await get_cached_stats(db)

# ---- Tractor Utilisation (weekly telematics CSV, uploaded by a manager) ----

@app.post("/api/tractor-utilisation/upload")
async def upload_tractor_utilisation(file: UploadFile = File(...)):
    """Parse and save the weekly tractor utilisation CSV. Columns are read by
    NAME (order doesn't matter): Nickname, Model, Idle (h), Working (h),
    Transport (h), Total Hours, Report End Date."""
    import csv as _csv
    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = _csv.DictReader(io.StringIO(text))

    def num(v):
        try:
            return float(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return 0.0

    rows = []
    report_end = ""
    for r in reader:
        nick = (r.get("Nickname") or "").strip()
        if not nick or nick == "---":
            continue
        idle = num(r.get("Idle (h)"))
        work = num(r.get("Working (h)"))
        trans = num(r.get("Transport (h)"))
        total = num(r.get("Total Hours"))
        if idle == 0 and work == 0 and trans == 0 and total == 0:
            continue
        if not report_end:
            report_end = (r.get("Report End Date") or "").strip()
        rows.append({
            "nickname": nick,
            "model": (r.get("Model") or "").strip(),
            "idle_h": round(idle, 1),
            "working_h": round(work, 1),
            "transport_h": round(trans, 1),
            "total_h": round(total, 1),
        })
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No machine rows found — check this is the weekly utilisation CSV "
                   "with columns: Nickname, Model, Idle (h), Working (h), Transport (h), Total Hours",
        )
    rows.sort(key=lambda x: -x["total_h"])
    doc = {
        "rows": rows,
        "report_end_date": report_end,
        "machine_count": len(rows),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.tractor_utilisation.delete_many({})
    await db.tractor_utilisation.insert_one({**doc})
    return doc

@app.get("/api/tractor-utilisation")
async def get_tractor_utilisation():
    """The latest saved tractor utilisation report."""
    doc = await db.tractor_utilisation.find_one({}, {"_id": 0})
    return doc or {"rows": [], "report_end_date": None, "machine_count": 0, "uploaded_at": None}

@app.get("/api/farm/crop-areas")
async def get_farm_crop_areas(year: int = 2027):
    """Our crop areas for a given year, parsed from the FieldPlan app's
    'Our crop areas — <year>' section (the local copy synced daily)."""
    import re as _re
    if not os.path.exists(FIELDPLAN_PATH):
        try:
            await download_fieldplan()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"FieldPlan not available: {str(e)}")
    with open(FIELDPLAN_PATH, "r", encoding="utf-8", errors="replace") as f:
        html = f.read()
    # Locate the section heading (em dash or hyphen between title and year)
    m = _re.search(rf"Our crop areas\s*[—-]\s*{year}\s*</h2>", html)
    if not m:
        raise HTTPException(status_code=404, detail=f"No crop areas found for {year}")
    seg = html[m.end():]
    # Stop at the next section heading, and before the partner-farmed list
    for marker in ('<div class="sh"', 'Partner-farmed'):
        end = seg.find(marker)
        if end != -1:
            seg = seg[:end]
    crops = []
    for cm in _re.finditer(
        r'background:\s*(#[0-9A-Fa-f]{3,6})"></div>'
        r'<span[^>]*>([^<]+)</span>'
        r'<span[^>]*>([\d,\.]+)\s*ha</span>',
        seg,
    ):
        color, name, ha = cm.group(1), cm.group(2).strip(), cm.group(3).replace(",", "")
        try:
            ha_val = float(ha)
        except ValueError:
            continue
        crops.append({"name": name, "ha": round(ha_val, 1), "color": color})
    crops.sort(key=lambda c: -c["ha"])
    return {
        "year": year,
        "crops": crops,
        "total_ha": round(sum(c["ha"] for c in crops), 1),
    }

# ---- Link to the Abreys Stock Control app (packouttracks) ----
STOCK_API_BASE = os.environ.get(
    "STOCK_API_BASE", "https://packouttracks-r-1774892359.emergent.host/api"
)

@app.get("/api/stock/summary")
async def get_stock_summary():
    """Live summary pulled from the Abreys Stock Control app:
    store utilisation per shed (grouped by crop) and grader throughput."""
    try:
        async with httpx.AsyncClient(timeout=20, follow_redirects=True) as hc:
            sheds_r, zones_r = await asyncio.gather(
                hc.get(f"{STOCK_API_BASE}/sheds"),
                hc.get(f"{STOCK_API_BASE}/zones"),
            )
            sheds_r.raise_for_status()
            zones_r.raise_for_status()
            sheds = sheds_r.json()
            zones = zones_r.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Stock app unreachable: {str(e)}")

    stores = []
    for shed in sheds:
        shed_zones = [z for z in zones if z.get("shed_id") == shed.get("id")]
        total = sum((z.get("total_quantity") or 0) for z in shed_zones)
        occupied = len([z for z in shed_zones if (z.get("total_quantity") or 0) > 0])
        utilization = round(occupied / len(shed_zones) * 100) if shed_zones else 0
        stores.append({
            "name": shed.get("name") or "",
            "crop_type": shed.get("crop_type") or "",
            "zones": len(shed_zones),
            "occupied_zones": occupied,
            "total_stock": round(total, 1),
            "utilization": utilization,
        })

    # Grader stats — the same feed the stock app's Lines Overview page uses:
    # per grader, the current session (T/H, tonnes in/out, waste, efficiency,
    # staff, hours) and all-time statistics.
    graders = []
    try:
        async with httpx.AsyncClient(timeout=25, follow_redirects=True) as hc:
            g_r = await hc.get(f"{STOCK_API_BASE}/grader-stats")
            if g_r.status_code == 200 and isinstance(g_r.json(), list):
                graders = g_r.json()
    except Exception:
        graders = []

    return {"stores": stores, "graders": graders}

@app.get("/api/dashboard/checks-by-day")
async def get_checks_by_day(days: int = 6):
    """Counts of completed checks per check type per day, for the last <days>
    calendar days (UK time, today included as the last day). Used by the
    dashboard's Check Figures section."""
    from zoneinfo import ZoneInfo
    uk = ZoneInfo("Europe/London")
    days = max(1, min(days, 14))
    today_uk = datetime.now(uk).date()
    day_list = [today_uk - timedelta(days=i) for i in range(days - 1, -1, -1)]
    start_utc = datetime.combine(day_list[0], datetime.min.time(), tzinfo=uk).astimezone(timezone.utc)

    # Pseudo-checklists that aren't real machine checks
    excluded = {"MACHINE ADD", "NEW MACHINE"}

    counts = {}
    day_totals = {d.isoformat(): 0 for d in day_list}
    cursor = db.checklists.find(
        {"completed_at": {"$gte": start_utc}},
        {"_id": 0, "completed_at": 1, "check_type": 1},
    )
    async for doc in cursor:
        completed = doc.get("completed_at")
        if isinstance(completed, str):
            try:
                completed = datetime.fromisoformat(completed.replace("Z", "+00:00"))
            except ValueError:
                continue
        if completed is None:
            continue
        if completed.tzinfo is None:
            completed = completed.replace(tzinfo=timezone.utc)
        day_key = completed.astimezone(uk).date().isoformat()
        if day_key not in day_totals:
            continue
        ct = doc.get("check_type") or "Unknown"
        if isinstance(ct, dict):
            ct = ct.get("check_type") or "Unknown"
        ct = str(ct).strip()
        if ct.upper() in excluded:
            continue
        counts.setdefault(ct, {})
        counts[ct][day_key] = counts[ct].get(day_key, 0) + 1
        day_totals[day_key] += 1

    today_key = today_uk.isoformat()
    today_by_type = {ct: per_day[today_key] for ct, per_day in counts.items() if per_day.get(today_key)}
    return {
        "days": [
            {"date": d.isoformat(), "label": d.strftime("%a %d %b"), "is_today": d == today_uk}
            for d in day_list
        ],
        "types": sorted(counts.keys(), key=lambda t: -sum(counts[t].values())),
        "counts": counts,
        "day_totals": day_totals,
        "today_by_type": today_by_type,
        "today_total": day_totals.get(today_key, 0),
    }

@app.get("/api/checklists", response_model=List[ChecklistResponse])
async def get_checklists(limit: int = 100, skip: int = 0, check_type: str = None):
    """Get checklists with pagination - optimized for speed"""
    # Build query filter
    query = {}
    if check_type:
        if ',' in check_type:
            check_types = [ct.strip() for ct in check_type.split(',')]
            query["check_type"] = {"$in": check_types}
        else:
            query["check_type"] = check_type
    
    # Enforce reasonable limits
    limit = min(limit, 500)  # Max 500 at a time
    
    try:
        checklists = await db.checklists.find(query, {"_id": 0}).sort("completed_at", -1).skip(skip).limit(limit).to_list(length=limit)
        
        # Parse datetime strings - simplified
        for checklist in checklists:
            if checklist.get('completed_at') and isinstance(checklist['completed_at'], str):
                try:
                    checklist['completed_at'] = datetime.fromisoformat(checklist['completed_at'].replace('Z', '+00:00'))
                except:
                    pass  # Keep as string if parsing fails
        
        return checklists
    except Exception as e:
        print(f"Error in get_checklists: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/checklists/today")
async def get_todays_checklists():
    """Get today's checklists - fast dedicated endpoint"""
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Use regex to match today's date regardless of time format
    checklists = await db.checklists.find(
        {"completed_at": {"$regex": f"^{today}"}},
        {"_id": 0}
    ).sort("completed_at", -1).to_list(length=100)
    
    # Parse datetime strings
    for checklist in checklists:
        if checklist.get('completed_at') and isinstance(checklist['completed_at'], str):
            try:
                checklist['completed_at'] = datetime.fromisoformat(checklist['completed_at'].replace('Z', '+00:00'))
            except:
                pass
    
    return checklists

@app.get("/api/checklists/by-machine")
async def get_checklists_by_machine(make: str = None, name: str = None, limit: int = 100, skip: int = 0):
    """Get checklists for a specific machine with pagination for better performance"""
    query = {}
    if make:
        query["machine_make"] = make
    if name:
        query["machine_model"] = name
    
    # Use projection to only get needed fields (faster)
    projection = {
        "_id": 0,
        "id": 1,
        "staff_name": 1,
        "machine_make": 1,
        "machine_model": 1,
        "check_type": 1,
        "completed_at": 1,
        "status": 1,
        "items_satisfactory": 1,
        "items_unsatisfactory": 1,
        "items_total": 1,
        "notes_summary": 1,
        "checklist_items": 1,
        "workshop_notes": 1,
        "workshop_photos": 1,
        # Fuel and Mileage fields
        "fuel_mileage": 1,
        "fuel_added": 1,
        "adblue_added": 1,
        "fuel_notes": 1
    }
    
    checklists = await db.checklists.find(query, projection).sort("completed_at", -1).skip(skip).limit(limit).to_list(length=limit)
    
    # Get total count for pagination info
    total = await db.checklists.count_documents(query)
    
    return {
        "checklists": checklists,
        "total": total,
        "limit": limit,
        "skip": skip,
        "has_more": skip + len(checklists) < total
    }

@app.get("/api/checklists/{checklist_id}", response_model=ChecklistResponse)
async def get_checklist(checklist_id: str):
    checklist = await db.checklists.find_one({"id": checklist_id}, {"_id": 0})
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    
    # Parse datetime string back to datetime object
    if isinstance(checklist['completed_at'], str):
        checklist['completed_at'] = datetime.fromisoformat(checklist['completed_at'])
    
    return ChecklistResponse(**checklist)

@app.get("/api/checklists-with-repairs")
async def get_checklists_with_repairs(limit: int = 50, skip: int = 0):
    """Get checklists that have unsatisfactory items OR are GENERAL REPAIR records"""
    # Build query to get checklists with unsatisfactory items or GENERAL REPAIR
    query = {
        "$or": [
            {"check_type": "GENERAL REPAIR"},
            {"checklist_items.status": "unsatisfactory"}
        ]
    }
    
    checklists = await db.checklists.find(query, {"_id": 0}).sort("completed_at", -1).skip(skip).limit(limit).to_list(length=limit)
    
    # Parse datetime strings
    for checklist in checklists:
        if isinstance(checklist.get('completed_at'), str):
            checklist['completed_at'] = datetime.fromisoformat(checklist['completed_at'])
    
    return checklists

@app.post("/api/admin/update-staff")
async def update_staff_list(staff_names: List[str]):
    """Update the staff list by replacing all existing staff with new list"""
    try:
        # Clear existing staff except admin (4444)
        await db.staff.delete_many({"employee_number": {"$ne": "4444"}})
        
        # Add new staff
        new_staff = []
        for staff_name in staff_names:
            staff = Staff(name=staff_name.strip())
            new_staff.append(staff.dict())
        
        if new_staff:
            await db.staff.insert_many(new_staff)
        
        return {"message": f"Successfully updated {len(new_staff)} staff members", "count": len(new_staff)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update staff list: {str(e)}")

class AssetUpdate(BaseModel):
    make: str
    model: str

@app.post("/api/admin/update-assets")
async def update_asset_list(assets: List[AssetUpdate]):
    """Update the asset list by replacing all existing assets with new list"""
    try:
        # Clear existing assets
        await db.assets.delete_many({})
        
        # Add new assets
        new_assets = []
        for asset_data in assets:
            asset = Asset(make=asset_data.make.strip(), model=asset_data.model.strip())
            new_assets.append(asset.dict())
        
        if new_assets:
            await db.assets.insert_many(new_assets)
        
        return {"message": f"Successfully updated {len(new_assets)} assets", "count": len(new_assets)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update asset list: {str(e)}")

# ============ OLD SharePoint OAuth Endpoints (Deprecated) ============
# These endpoints use user OAuth flow which requires manual authentication.
# Now replaced by SharePoint Auto-Sync with client credentials (app-only auth).
# Keeping them commented for reference but they won't work in production.

# @app.get("/api/admin/sharepoint/auth-url")
# @app.post("/api/admin/sharepoint/callback")
# @app.get("/api/admin/sharepoint/test")
# @app.post("/api/admin/sharepoint/sync-staff")

class AuthCallbackRequest(BaseModel):
    auth_code: str

# OLD SharePoint sync endpoints removed - now using SharePoint Auto-Sync with client credentials

@app.post("/api/admin/upload-staff-file")
async def upload_staff_file(file: UploadFile = File(...)):
    """Upload and process staff with employee numbers from Excel file"""
    try:
        import openpyxl
        from io import BytesIO
        
        # Read file content
        file_content = await file.read()
        print(f"[STAFF UPLOAD] File received: {file.filename}, size: {len(file_content)} bytes")
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook[workbook.sheetnames[0]]  # Use first sheet, not active
        print(f"[STAFF UPLOAD] Sheet name: {workbook.sheetnames[0]}, max_row: {sheet.max_row}, max_col: {sheet.max_column}")
        
        # Get headers and find name/employee number/workshop control/admin control/manager control columns
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        print(f"[STAFF UPLOAD] Headers found: {headers}")
        
        name_col = None
        number_col = None
        workshop_col = None
        admin_col = None
        manager_col = None
        
        for i, header in enumerate(headers):
            # Check for employee number column FIRST (more specific match)
            if ('employee' in header and 'number' in header) or header == 'emp no' or header == 'emp number' or header == 'employee_number':
                number_col = i
            elif 'name' in header and 'employee' not in header:
                name_col = i
            elif 'workshop' in header and 'control' in header:
                workshop_col = i
            elif 'admin' in header and 'control' in header:
                admin_col = i
            elif 'manager' in header:  # Accept "Manager" or "Manager Control"
                manager_col = i
        
        # If we didn't find employee number yet, look for other patterns (but NOT phone number)
        if number_col is None:
            for i, header in enumerate(headers):
                if ('number' in header or 'emp' in header) and 'phone' not in header and 'tel' not in header and 'mob' not in header:
                    number_col = i
                    break
        
        print(f"[STAFF UPLOAD] Column mapping - name_col: {name_col}, number_col: {number_col}, workshop_col: {workshop_col}, admin_col: {admin_col}, manager_col: {manager_col}")
        
        # Fallback: assume first column is names, second is numbers
        if name_col is None:
            name_col = 0
            print(f"[STAFF UPLOAD] Using fallback name_col: 0")
        if number_col is None and len(headers) > 1:
            number_col = 1
            print(f"[STAFF UPLOAD] Using fallback number_col: 1")
        
        if number_col is None:
            raise HTTPException(status_code=400, detail="Could not find Employee Number column. Please ensure your Excel has both Name and Employee Number columns.")
        
        # Extract staff data
        staff_data = []
        rows_processed = 0
        rows_skipped = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            rows_processed += 1
            if row and len(row) > max(name_col, number_col):
                name = str(row[name_col]).strip() if row[name_col] else ''
                emp_number = str(row[number_col]).strip() if row[number_col] else ''
                workshop_control = None
                admin_control = None
                manager_control = None
                
                if workshop_col is not None and len(row) > workshop_col and row[workshop_col]:
                    workshop_control = str(row[workshop_col]).strip().lower()
                
                if admin_col is not None and len(row) > admin_col and row[admin_col]:
                    admin_control = str(row[admin_col]).strip().lower()
                
                if manager_col is not None and len(row) > manager_col and row[manager_col]:
                    manager_control = str(row[manager_col]).strip().lower()
                
                if name and emp_number and name.lower() not in ['name', 'staff', 'employee']:
                    staff_data.append({
                        "name": name,
                        "employee_number": emp_number,
                        "active": True,
                        "workshop_control": workshop_control,
                        "admin_control": admin_control,
                        "manager_control": manager_control
                    })
                else:
                    rows_skipped += 1
                    if rows_processed <= 5:
                        print(f"[STAFF UPLOAD] Skipped row {rows_processed}: name='{name}', emp_number='{emp_number}'")
            else:
                rows_skipped += 1
        
        print(f"[STAFF UPLOAD] Rows processed: {rows_processed}, valid staff: {len(staff_data)}, skipped: {rows_skipped}")
        
        if not staff_data:
            raise HTTPException(status_code=400, detail=f"No valid staff data found. Processed {rows_processed} rows but none had valid Name and Employee Number. Headers found: {headers}")
        
        # Update database - preserve admin account (4444)
        delete_result = await db.staff.delete_many({"employee_number": {"$ne": "4444"}})
        print(f"[STAFF UPLOAD] Deleted {delete_result.deleted_count} existing staff records")
        
        new_staff = [Staff(**data).dict() for data in staff_data]
        insert_result = await db.staff.insert_many(new_staff)
        print(f"[STAFF UPLOAD] Inserted {len(insert_result.inserted_ids)} new staff records")
        
        return {
            "message": f"Successfully uploaded {len(staff_data)} staff members with employee numbers",
            "count": len(staff_data),
            "preview": staff_data[:5],
            "debug": {
                "headers_found": headers,
                "rows_processed": rows_processed,
                "rows_skipped": rows_skipped
            }
        }
        
    except Exception as e:
        import traceback
        print(f"[STAFF UPLOAD ERROR] {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to process staff file: {str(e)}")


# ============ SharePoint Auto-Sync Endpoints ============

@app.get("/api/admin/sharepoint/test-connection")
async def test_sharepoint_connection():
    """Test the SharePoint connection and return file info"""
    try:
        result = sharepoint_auto_sync.test_connection()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Connection test failed: {str(e)}")

@app.get("/api/admin/sharepoint/debug-env")
async def debug_sharepoint_env():
    """Debug endpoint to check if Azure environment variables are set (without revealing secrets)"""
    import os
    return {
        "AZURE_CLIENT_ID_SET": bool(os.environ.get('AZURE_CLIENT_ID')),
        "AZURE_CLIENT_ID_LENGTH": len(os.environ.get('AZURE_CLIENT_ID', '')),
        "AZURE_CLIENT_SECRET_SET": bool(os.environ.get('AZURE_CLIENT_SECRET')),
        "AZURE_CLIENT_SECRET_LENGTH": len(os.environ.get('AZURE_CLIENT_SECRET', '')),
        "AZURE_TENANT_ID_SET": bool(os.environ.get('AZURE_TENANT_ID')),
        "AZURE_TENANT_ID_LENGTH": len(os.environ.get('AZURE_TENANT_ID', '')),
        "SHAREPOINT_SITE_URL": os.environ.get('SHAREPOINT_SITE_URL', 'NOT SET'),
        "SHAREPOINT_STAFF_FILENAME": os.environ.get('SHAREPOINT_STAFF_FILENAME', 'NOT SET'),
        "SHAREPOINT_ASSETS_FILENAME": os.environ.get('SHAREPOINT_ASSETS_FILENAME', 'NOT SET')
    }

@app.post("/api/admin/sharepoint/sync-now")
async def trigger_sharepoint_sync():
    """Manually trigger a SharePoint staff sync"""
    try:
        result = await sharepoint_auto_sync.sync_staff_list(db)
        
        # Log the sync
        await db.sync_logs.insert_one({
            'type': 'manual_staff',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'success': result.get('success', False),
            'message': result.get('message', ''),
            'count': result.get('count', 0)
        })
        
        if result.get('success'):
            return result
        else:
            raise HTTPException(status_code=500, detail=result.get('message', 'Sync failed'))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

@app.post("/api/admin/sharepoint/sync-assets")
async def trigger_assets_sync():
    """Manually trigger a SharePoint assets sync"""
    try:
        result = await sharepoint_auto_sync.sync_assets_list(db)
        
        # Log the sync
        await db.sync_logs.insert_one({
            'type': 'manual_assets',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'success': result.get('success', False),
            'message': result.get('message', ''),
            'assets_count': result.get('assets_count', 0),
            'templates_count': result.get('templates_count', 0)
        })
        
        if result.get('success'):
            return result
        else:
            raise HTTPException(status_code=500, detail=result.get('message', 'Sync failed'))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

@app.post("/api/admin/sharepoint/sync-all")
async def trigger_full_sync():
    """Manually trigger a full SharePoint sync (staff + assets)"""
    try:
        result = await sharepoint_auto_sync.sync_all(db)
        
        # Log the sync
        await db.sync_logs.insert_one({
            'type': 'manual_all',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'success': result.get('success', False),
            'message': f"Staff: {result.get('staff', {}).get('message', 'N/A')}, Assets: {result.get('assets', {}).get('message', 'N/A')}",
            'staff_count': result.get('staff', {}).get('count', 0),
            'assets_count': result.get('assets', {}).get('assets_count', 0),
            'templates_count': result.get('assets', {}).get('templates_count', 0)
        })
        
        if result.get('success'):
            return result
        else:
            raise HTTPException(status_code=500, detail=result.get('message', 'Sync failed'))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

@app.get("/api/admin/sharepoint/sync-status")
async def get_sync_status():
    """Get the status of SharePoint sync including last sync time and next scheduled sync"""
    try:
        # Get last sync log
        last_sync = await db.sync_logs.find_one(
            {},
            sort=[('timestamp', -1)]
        )
        
        # Get scheduler job info
        job = scheduler.get_job('daily_staff_sync')
        next_run = job.next_run_time.isoformat() if job and job.next_run_time else None
        
        return {
            'last_sync': {
                'timestamp': last_sync.get('timestamp') if last_sync else None,
                'success': last_sync.get('success') if last_sync else None,
                'message': last_sync.get('message') if last_sync else None,
                'count': last_sync.get('count') if last_sync else None,
                'type': last_sync.get('type') if last_sync else None
            } if last_sync else None,
            'next_scheduled_sync': next_run,
            'scheduler_running': scheduler.running
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get sync status: {str(e)}")

@app.get("/api/admin/sharepoint/sync-logs")
async def get_sync_logs(limit: int = 10):
    """Get recent sync logs"""
    try:
        logs = await db.sync_logs.find(
            {},
            {'_id': 0}
        ).sort('timestamp', -1).limit(limit).to_list(length=limit)
        return logs
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get sync logs: {str(e)}")


@app.get("/api/admin/template-diagnostics")
async def get_template_diagnostics():
    """Diagnostic endpoint to verify template-to-asset mappings"""
    try:
        # Get all unique check_types from assets
        assets = await db.assets.find({}, {"_id": 0, "check_type": 1, "name": 1}).to_list(length=10000)
        check_type_counts = {}
        for a in assets:
            ct = a.get('check_type', 'unknown')
            check_type_counts[ct] = check_type_counts.get(ct, 0) + 1
        
        # Get all templates
        templates = await db.checklist_templates.find({}, {"_id": 0}).to_list(length=100)
        template_info = []
        for t in templates:
            items = t.get('items', [])
            template_info.append({
                'check_type': t.get('check_type'),
                'sheet_name': t.get('sheet_name'),
                'item_count': len(items),
                'first_3_items': [i.get('item', '') for i in items[:3]],
                'updated_at': t.get('updated_at'),
                'assets_using_this': check_type_counts.get(t.get('check_type'), 0)
            })
        
        # Find check_types with no template
        template_types = set(t.get('check_type') for t in templates)
        missing_templates = [ct for ct in check_type_counts if ct not in template_types]
        
        return {
            'total_assets': len(assets),
            'check_type_counts': check_type_counts,
            'templates': template_info,
            'missing_templates': missing_templates
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get diagnostics: {str(e)}")



@app.post("/api/admin/upload-assets-file") 
async def upload_assets_file(file: UploadFile = File(...)):
    """Upload and process assets from Excel file"""
    try:
        import openpyxl
        from io import BytesIO
        
        # Read file content
        file_content = await file.read()
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook[workbook.sheetnames[0]]  # Use first sheet, not active
        
        # Get headers and find check_type, name, make columns
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        check_type_col = None
        name_col = None
        make_col = None
        
        for i, header in enumerate(headers):
            if header == 'check type' or 'checktype' in header:
                check_type_col = i
            elif header == 'namecolumn' or 'name' in header:
                name_col = i
            elif header == 'makecolumn' or 'make' in header:
                make_col = i
        
        if check_type_col is None or name_col is None or make_col is None:
            raise HTTPException(status_code=400, detail="Could not find Check Type, Name of Implement, and Make columns in the file")
        
        # Extract asset data
        assets = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if row and len(row) > max(check_type_col, name_col, make_col):
                check_type = str(row[check_type_col]).strip() if row[check_type_col] else ''
                name = str(row[name_col]).strip() if row[name_col] else ''
                make = str(row[make_col]).strip() if row[make_col] else ''
                
                if check_type and name and make:
                    assets.append({
                        "check_type": check_type,
                        "name": name, 
                        "make": make
                    })
        
        if not assets:
            raise HTTPException(status_code=400, detail="No asset data found in the uploaded file")
        
        # Get existing assets to preserve QR print status
        existing_assets = await db.assets.find({}, {"_id": 0}).to_list(length=10000)
        existing_qr_status = {}
        for ea in existing_assets:
            # Key by make+name to match assets
            key = f"{ea.get('make', '')}:{ea.get('name', '')}"
            if ea.get('qr_printed'):
                existing_qr_status[key] = {
                    'qr_printed': ea.get('qr_printed', False),
                    'qr_printed_at': ea.get('qr_printed_at')
                }
        
        # Update assets database - preserve QR status for existing machines
        await db.assets.delete_many({})
        new_assets = []
        for asset in assets:
            asset_obj = Asset(**asset)
            asset_dict = asset_obj.dict()
            # Check if this asset had QR printed before
            key = f"{asset_dict['make']}:{asset_dict['name']}"
            if key in existing_qr_status:
                asset_dict['qr_printed'] = existing_qr_status[key]['qr_printed']
                asset_dict['qr_printed_at'] = existing_qr_status[key]['qr_printed_at']
            new_assets.append(asset_dict)
        await db.assets.insert_many(new_assets)
        
        # Process checklist sheets
        checklist_templates = []
        processed_sheets = []
        
        # Get all unique check types from assets
        unique_check_types = set(asset['check_type'] for asset in assets)
        
        # Process each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Skip the main asset sheet (first sheet)
            if sheet_name == workbook.sheetnames[0]:
                continue
            
            # Try to match sheet name with check types - improved matching
            matching_check_type = None
            sheet_name_clean = sheet_name.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '').replace('checklist', '')
            
            # First try exact matches
            for check_type in unique_check_types:
                check_type_clean = check_type.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '').replace('checklist', '')
                if sheet_name_clean == check_type_clean or check_type.lower() == sheet_name.lower():
                    matching_check_type = check_type
                    break
            
            # If no exact match, try partial matches
            if not matching_check_type:
                for check_type in unique_check_types:
                    check_type_clean = check_type.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '').replace('checklist', '')
                    if sheet_name_clean in check_type_clean or check_type_clean in sheet_name_clean:
                        matching_check_type = check_type
                        break
            
            # If still no match, use sheet name as check type
            if not matching_check_type:
                matching_check_type = sheet_name
            
            # Extract checklist items from this sheet
            # First, find the header row and locate the "Compulsory" column
            items = []
            compulsory_col = None
            item_col = 0  # Default to first column for item text
            
            # Get headers from first row to find Compulsory column
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            if header_row:
                for col_idx, header in enumerate(header_row):
                    if header:
                        header_lower = str(header).strip().lower()
                        if 'compulsory' in header_lower or 'compulsary' in header_lower:  # Handle common misspelling
                            compulsory_col = col_idx
                        elif 'item' in header_lower or 'task' in header_lower or 'check' in header_lower or 'description' in header_lower:
                            item_col = col_idx
            
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_num == 1:  # Skip header row
                    continue
                    
                if row and len(row) > item_col and row[item_col]:  # If item column has content
                    item_text = str(row[item_col]).strip()
                    # Skip obvious headers or empty items
                    if (item_text and 
                        item_text.lower() not in ['item', 'check', 'description', 'checklist', 'safety'] and
                        len(item_text) > 3):  # Minimum length filter
                        
                        # Check if item is marked as compulsory
                        is_compulsory = False
                        if compulsory_col is not None and len(row) > compulsory_col and row[compulsory_col]:
                            compulsory_value = str(row[compulsory_col]).strip().lower()
                            is_compulsory = compulsory_value in ['yes', 'y', 'true', '1', 'x', 'compulsory']
                        
                        items.append({"item": item_text, "compulsory": is_compulsory})
            
            if items:
                compulsory_count = sum(1 for item in items if item.get('compulsory', False))
                template = {
                    "id": str(uuid.uuid4()),
                    "check_type": matching_check_type,
                    "items": items
                }
                checklist_templates.append(template)
                processed_sheets.append(f"{sheet_name} -> {matching_check_type} ({len(items)} items, {compulsory_count} compulsory)")
        
        # Update checklist templates in database
        if checklist_templates:
            # Clear ALL existing templates and insert new ones for complete refresh
            await db.checklist_templates.delete_many({})
            
            # Insert new templates
            await db.checklist_templates.insert_many(checklist_templates)
        
        return {
            "message": f"Successfully uploaded {len(assets)} assets and {len(checklist_templates)} checklist templates", 
            "count": len(assets),
            "templates_created": len(checklist_templates),
            "processed_sheets": processed_sheets,
            "preview": assets[:5] if assets else []
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process assets file: {str(e)}")

@app.post("/api/admin/upload-checklist-file/{check_type}")
async def upload_checklist_file(check_type: str, file: UploadFile = File(...)):
    """Upload and process checklist template from Excel file"""
    try:
        import openpyxl
        from io import BytesIO
        
        # Validate check type
        valid_types = ['daily_check', 'grader_startup', 'workshop_service']
        if check_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid check type. Must be one of: {valid_types}")
        
        # Read file content
        file_content = await file.read()
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook[workbook.sheetnames[0]]  # Use first sheet, not active
        
        # Get headers and find required columns
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        item_col = None
        category_col = None
        critical_col = None
        
        for i, header in enumerate(headers):
            if 'item' in header or 'task' in header:
                item_col = i
            elif 'category' in header:
                category_col = i
            elif 'critical' in header or 'common' in header:
                critical_col = i
        
        if item_col is None:
            raise HTTPException(status_code=400, detail="Could not find Item or Task column in the file")
        
        # Extract checklist items
        items = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if row and len(row) > item_col and row[item_col]:
                item_text = str(row[item_col]).strip()
                if item_text:
                    items.append(item_text)
        
        if not items:
            raise HTTPException(status_code=400, detail="No checklist items found in the uploaded file")
        
        # Update database
        await db.checklist_templates.delete_many({"check_type": check_type})
        
        template = ChecklistTemplate(
            check_type=check_type,
            items=items
        )
        await db.checklist_templates.insert_one(template.dict())
        
        return {
            "message": f"Successfully uploaded {len(items)} items for {check_type}",
            "count": len(items),
            "check_type": check_type,
            "preview": items[:5]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process checklist file: {str(e)}")

# OLD sync-checklists endpoint removed - now handled by sync_assets_list which processes
# checklist templates from the AssetList.xlsx file sheets

@app.get("/api/checklist-templates/{check_type:path}")
async def get_checklist_template(check_type: str):
    """Get checklist template for a specific check type"""
    try:
        template = await db.checklist_templates.find_one({"check_type": check_type}, {"_id": 0})
        if template:
            # Ensure items have compulsory flag (for backward compatibility)
            if template.get('items'):
                for i, item in enumerate(template['items']):
                    if isinstance(item, str):
                        # Convert old string format to new object format
                        template['items'][i] = {"item": item, "compulsory": False}
                    elif isinstance(item, dict) and 'compulsory' not in item:
                        item['compulsory'] = False
            return template
        else:
            # Return default templates if not found in database
            default_templates = {
                'daily_check': [
                    {"item": "Oil level check - Engine oil at correct level", "compulsory": False},
                    {"item": "Fuel level check - Adequate fuel for operation", "compulsory": False},
                    {"item": "Hydraulic fluid level - Within acceptable range", "compulsory": False},
                    {"item": "Battery condition - Terminals clean, voltage adequate", "compulsory": False},
                    {"item": "Tire/track condition - No visible damage or excessive wear", "compulsory": False},
                    {"item": "Safety guards in place - All protective covers secured", "compulsory": True},
                    {"item": "Emergency stop function - Test emergency stop button", "compulsory": True},
                    {"item": "Warning lights operational - All safety lights working", "compulsory": False},
                    {"item": "Operator seat condition - Seat belt and controls functional", "compulsory": False},
                    {"item": "Air filter condition - Clean and properly sealed", "compulsory": False},
                    {"item": "Cooling system - Radiator clear, coolant level adequate", "compulsory": False},
                    {"item": "Brake system function - Service and parking brakes operational", "compulsory": True},
                    {"item": "Steering operation - Smooth operation, no excessive play", "compulsory": False},
                    {"item": "Lights and signals - All operational lights working", "compulsory": False},
                    {"item": "Fire extinguisher - Present and within service date", "compulsory": True}
                ],
                'grader_startup': [
                    {"item": "Emergency stops working and present - Test all emergency stop buttons", "compulsory": True},
                    {"item": "Walkways clear of debris and gates closed - All access areas safe", "compulsory": True},
                    {"item": "Guards are all in place - All safety guards properly secured", "compulsory": True},
                    {"item": "All personnel accounted for and out of reach of dangers - Safety zone clear", "compulsory": True},
                    {"item": "Oil level check - Engine oil at correct level", "compulsory": False},
                    {"item": "Fuel level check - Adequate fuel for operation", "compulsory": False},
                    {"item": "Hydraulic fluid level - Within acceptable range", "compulsory": False},
                    {"item": "Battery condition - Terminals clean, voltage adequate", "compulsory": False},
                    {"item": "Track/blade condition - No visible damage or excessive wear", "compulsory": False},
                    {"item": "Blade operation - Hydraulic lift and angle functions working", "compulsory": False},
                    {"item": "Warning beacon - Rotating warning light operational", "compulsory": False},
                    {"item": "Backup alarm - Reverse warning signal functional", "compulsory": False}
                ],
                'workshop_service': []
            }
            
            items = default_templates.get(check_type, [])
            return {
                "check_type": check_type,
                "items": items,
                "source": "default"
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get checklist template: {str(e)}")

# OLD SharePoint sync endpoint removed - using new sharepoint_auto_sync with client credentials flow

@app.get("/api/checklists/export/csv")
async def export_checklists_csv():
    """Fast CSV export - use this for very large datasets"""
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    # Use projection for speed
    projection = {
        "_id": 0, "id": 1, "staff_name": 1, "machine_make": 1, "machine_model": 1,
        "check_type": 1, "completed_at": 1, "status": 1, "checklist_items": 1, "workshop_notes": 1
    }
    
    checklists = await db.checklists.find({}, projection).sort("completed_at", -1).limit(10000).to_list(length=10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(["ID", "Staff Name", "Machine Make", "Machine Model", "Check Type", "Completed At", "Status", "Satisfactory", "Unsatisfactory", "Total", "Notes", "Workshop Details"])
    
    # Write data
    for checklist in checklists:
        check_type = checklist.get('check_type', '')
        if check_type in ['daily_check', 'grader_startup']:
            items = checklist.get('checklist_items', [])
            items_satisfactory = sum(1 for item in items if item.get('status') == 'satisfactory')
            items_unsatisfactory = sum(1 for item in items if item.get('status') == 'unsatisfactory')
            items_total = len(items)
            notes_list = [item.get('notes', '')[:100] for item in items if item.get('notes')]
            notes = "; ".join(notes_list)[:500] if notes_list else ""
            workshop_details = ""
        else:
            items_satisfactory = 0
            items_unsatisfactory = 0
            items_total = 0
            notes = ""
            workshop_details = (checklist.get('workshop_notes') or '')[:500]
        
        writer.writerow([
            checklist.get('id', ''),
            checklist.get('staff_name', ''),
            checklist.get('machine_make', ''),
            checklist.get('machine_model', ''),
            check_type,
            checklist.get('completed_at', ''),
            checklist.get('status', ''),
            items_satisfactory,
            items_unsatisfactory,
            items_total,
            notes,
            workshop_details
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/csv',
        headers={"Content-Disposition": "attachment; filename=all_checks.csv"}
    )

@app.get("/api/checklists/export/excel")
async def export_checklists_excel():
    """Optimized Excel export for large datasets"""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Use projection to only get fields we need (reduces memory)
    projection = {
        "_id": 0, "id": 1, "staff_name": 1, "machine_make": 1, "machine_model": 1,
        "check_type": 1, "completed_at": 1, "status": 1, "checklist_items": 1, "workshop_notes": 1
    }
    
    # Stream data in batches to avoid memory issues
    checklists = await db.checklists.find({}, projection).sort("completed_at", -1).limit(10000).to_list(length=10000)
    
    # Create workbook with optimized settings
    wb = Workbook(write_only=False)  # Can't use write_only with formatting
    ws = wb.active
    ws.title = "All Checks"
    
    # Define headers and fixed column widths (skip auto-adjust which is slow)
    headers = ["ID", "Staff Name", "Machine Make", "Machine Model", "Check Type", "Completed At", "Status", "Satisfactory", "Unsatisfactory", "Total", "Notes", "Workshop Details"]
    col_widths = [38, 20, 20, 25, 15, 22, 12, 12, 14, 8, 50, 50]
    
    # Set column widths upfront (much faster than auto-adjust)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Write and format header
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Process data in optimized way
    for checklist in checklists:
        check_type = checklist.get('check_type', '')
        
        if check_type in ['daily_check', 'grader_startup']:
            items = checklist.get('checklist_items', [])
            items_satisfactory = sum(1 for item in items if item.get('status') == 'satisfactory')
            items_unsatisfactory = sum(1 for item in items if item.get('status') == 'unsatisfactory')
            items_total = len(items)
            # Limit notes length to prevent huge cells
            notes_list = [item.get('notes', '')[:100] for item in items if item.get('notes')]
            notes = "; ".join(notes_list)[:500] if notes_list else ""
            workshop_details = ""
        else:
            items_satisfactory = 0
            items_unsatisfactory = 0
            items_total = 0
            notes = ""
            workshop_details = (checklist.get('workshop_notes') or '')[:500]
        
        ws.append([
            checklist.get('id', ''),
            checklist.get('staff_name', ''),
            checklist.get('machine_make', ''),
            checklist.get('machine_model', ''),
            check_type,
            str(checklist.get('completed_at', '')),
            checklist.get('status', ''),
            items_satisfactory,
            items_unsatisfactory,
            items_total,
            notes,
            workshop_details
        ])
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=all_checks.xlsx"}
    )

@app.get("/api/checklists/export/excel-by-machine")
async def export_checklists_excel_by_machine(make: str = None, name: str = None):
    """Export checklists to Excel - optimized for large datasets.
    Uses CSV-style approach for speed, then converts to Excel."""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Build query
    query = {}
    if make:
        query["machine_make"] = make
    if name:
        query["machine_model"] = name
    
    # Use projection to only get needed fields - MUCH faster
    projection = {
        "_id": 0,
        "id": 1,
        "staff_name": 1,
        "machine_make": 1,
        "machine_model": 1,
        "check_type": 1,
        "completed_at": 1,
        "checklist_items": 1,
        "workshop_notes": 1,
        "notes_summary": 1,
        "items_satisfactory": 1,
        "items_unsatisfactory": 1,
        "items_total": 1
    }
    
    # Stream results in batches for memory efficiency
    checklists = []
    cursor = db.checklists.find(query, projection).sort("completed_at", -1)
    async for doc in cursor:
        checklists.append(doc)
        if len(checklists) >= 10000:  # Cap at 10k for reasonable export time
            break
    
    if not checklists:
        raise HTTPException(status_code=404, detail="No checklists found")
    
    # Group by check_type
    by_type = {}
    for c in checklists:
        ct = c.get('check_type', 'unknown')
        if ct not in by_type:
            by_type[ct] = []
        by_type[ct].append(c)
    
    # Get templates (for question columns)
    templates = {}
    async for t in db.checklist_templates.find({}, {"_id": 0, "check_type": 1, "items": 1}):
        templates[t.get('check_type')] = t.get('items', [])
    
    # Create workbook
    wb = Workbook(write_only=False)  # write_only=True is faster but limited
    
    # Simple styles (apply sparingly for speed)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    first_sheet = True
    
    for check_type, type_checklists in sorted(by_type.items()):
        if not type_checklists:
            continue
        
        # Create sheet
        sheet_name = check_type[:31].replace('/', '-').replace('\\', '-')
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # Get all question items for this type
        all_items = []
        if check_type in templates:
            for item in templates[check_type]:
                if isinstance(item, dict):
                    all_items.append(item.get('item', ''))
                else:
                    all_items.append(str(item))
        
        # Also collect from actual data
        for c in type_checklists[:50]:  # Sample first 50 for speed
            for item in c.get('checklist_items', []):
                item_name = item.get('item', '')
                if item_name and item_name not in all_items:
                    all_items.append(item_name)
        
        # Headers
        if check_type == 'workshop_service' or not all_items:
            headers = ["Date", "Time", "Staff", "Machine Make", "Machine Model", "Notes"]
        else:
            headers = ["Date", "Time", "Staff", "Machine Make", "Machine Model"] + all_items[:50] + ["Notes"]  # Limit columns
        
        # Write header row with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header[:50])  # Truncate long headers
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data rows - minimal styling for speed
        for row_idx, c in enumerate(type_checklists, 2):
            completed = c.get('completed_at', '')
            if isinstance(completed, str) and len(completed) >= 10:
                date_str = completed[:10]
                time_str = completed[11:16] if len(completed) > 16 else ''
            else:
                date_str = str(completed)[:10] if completed else ''
                time_str = ''
            
            row = [
                date_str,
                time_str,
                c.get('staff_name', ''),
                c.get('machine_make', ''),
                c.get('machine_model', '')
            ]
            
            if check_type == 'workshop_service' or not all_items:
                row.append(c.get('workshop_notes', '') or c.get('notes_summary', ''))
            else:
                # Build status map
                status_map = {}
                notes = []
                for item in c.get('checklist_items', []):
                    status_map[item.get('item', '')] = item.get('status', '')
                    if item.get('notes'):
                        notes.append(item['notes'][:30])
                
                # Add status for each question column
                for item_name in all_items[:50]:
                    status = status_map.get(item_name, '')
                    if status == 'satisfactory':
                        row.append('✓')
                    elif status == 'unsatisfactory':
                        row.append('✗')
                    elif status == 'n/a':
                        row.append('N/A')
                    else:
                        row.append('')
                
                row.append('; '.join(notes) if notes else c.get('notes_summary', ''))
            
            # Write row (no cell-by-cell styling for speed)
            for col, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Set column widths (do this once, not per cell)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    summary.append(["Check Type", "Count"])
    summary.cell(row=1, column=1).fill = header_fill
    summary.cell(row=1, column=1).font = header_font
    summary.cell(row=1, column=2).fill = header_fill
    summary.cell(row=1, column=2).font = header_font
    
    for ct, cl in sorted(by_type.items()):
        summary.append([ct, len(cl)])
    
    summary.append([])
    summary.append(["Total", len(checklists)])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=checklists_export.xlsx"}
    )

# Repair Status Management Endpoints
class RepairStatusUpdate(BaseModel):
    repair_id: str
    acknowledged: Optional[bool] = None
    completed: Optional[bool] = None
    progress_notes: Optional[List[dict]] = None

@app.get("/api/repair-status/bulk")
async def get_bulk_repair_status():
    """Get status for all repairs"""
    statuses = await db.repair_status.find({}, {"_id": 0}).to_list(length=10000)  # Max 10000 statuses
    # Return as a dictionary keyed by repair_id for easy lookup
    return {status["repair_id"]: status for status in statuses}

@app.get("/api/repair-status/{repair_id}")
async def get_repair_status(repair_id: str):
    """Get status of a specific repair"""
    status = await db.repair_status.find_one({"repair_id": repair_id}, {"_id": 0})
    if not status:
        return {"repair_id": repair_id, "acknowledged": False, "completed": False, "progress_notes": []}
    return status

@app.post("/api/repair-status/acknowledge")
async def acknowledge_repair(repair_id: str):
    """Mark a repair as acknowledged"""
    await db.repair_status.update_one(
        {"repair_id": repair_id},
        {"$set": {
            "repair_id": repair_id,
            "acknowledged": True,
            "acknowledged_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    # Invalidate dashboard cache so counts update immediately
    await invalidate_cache()
    return {"success": True, "message": "Repair acknowledged"}

@app.post("/api/repair-status/complete")
async def complete_repair(repair_id: str):
    """Mark a repair as completed"""
    await db.repair_status.update_one(
        {"repair_id": repair_id},
        {"$set": {
            "completed": True,
            "completed_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    # Invalidate dashboard cache so counts update immediately
    await invalidate_cache()
    return {"success": True, "message": "Repair marked as complete"}

@app.post("/api/repair-status/add-note")
async def add_progress_note(repair_id: str, note_text: str, author: str):
    """Add a progress note to a repair"""
    note = {
        "text": note_text,
        "author": author,
        "date": datetime.now(timezone.utc).isoformat()
    }
    
    await db.repair_status.update_one(
        {"repair_id": repair_id},
        {
            "$push": {"progress_notes": note},
            "$setOnInsert": {"repair_id": repair_id, "acknowledged": False, "completed": False}
        },
        upsert=True
    )
    return {"success": True, "message": "Progress note added", "note": note}

# ============================================
# Near Miss and Suggestion Endpoints
# ============================================

@app.post("/api/near-misses")
async def create_near_miss(near_miss: NearMissCreate, employee_number: str = None):
    """Submit a new near miss report"""
    near_miss_doc = {
        "id": str(uuid.uuid4()),
        "description": near_miss.description,
        "location": near_miss.location,
        "photos": near_miss.photos,
        "is_anonymous": near_miss.is_anonymous,
        "submitted_by": near_miss.submitted_by if not near_miss.is_anonymous else None,
        "employee_number": employee_number if not near_miss.is_anonymous else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "acknowledged": False
    }
    await db.near_misses.insert_one(near_miss_doc)
    await invalidate_cache()
    return {"success": True, "message": "Near miss reported successfully", "id": near_miss_doc["id"]}

@app.get("/api/near-misses")
async def get_near_misses(acknowledged: bool = None, limit: int = 100):
    """Get near miss reports"""
    query = {}
    if acknowledged is not None:
        query["acknowledged"] = acknowledged
    
    near_misses = await db.near_misses.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return near_misses

@app.get("/api/near-misses/count")
async def get_near_misses_count():
    """Get count of new (unacknowledged) near misses"""
    total = await db.near_misses.count_documents({})
    new_count = await db.near_misses.count_documents({"acknowledged": False})
    return {"total": total, "new": new_count}

@app.post("/api/near-misses/{near_miss_id}/acknowledge")
async def acknowledge_near_miss(near_miss_id: str, acknowledged_by: str = "Admin"):
    """Acknowledge a near miss report"""
    result = await db.near_misses.update_one(
        {"id": near_miss_id},
        {"$set": {
            "acknowledged": True,
            "acknowledged_at": datetime.now(timezone.utc).isoformat(),
            "acknowledged_by": acknowledged_by
        }}
    )
    await invalidate_cache()
    if result.modified_count > 0:
        return {"success": True, "message": "Near miss acknowledged"}
    raise HTTPException(status_code=404, detail="Near miss not found")

@app.post("/api/suggestions")
async def create_suggestion(suggestion: SuggestionCreate, employee_number: str = None):
    """Submit a new suggestion"""
    suggestion_doc = {
        "id": str(uuid.uuid4()),
        "title": suggestion.title,
        "description": suggestion.description,
        "category": suggestion.category,
        "location": suggestion.location,
        "photos": suggestion.photos,
        "is_anonymous": suggestion.is_anonymous,
        "submitted_by": suggestion.submitted_by if not suggestion.is_anonymous else None,
        "employee_number": employee_number if not suggestion.is_anonymous else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "new"
    }
    await db.suggestions.insert_one(suggestion_doc)
    await invalidate_cache()
    return {"success": True, "message": "Suggestion submitted successfully", "id": suggestion_doc["id"]}

@app.get("/api/suggestions")
async def get_suggestions(status: str = None, limit: int = 100):
    """Get suggestions"""
    query = {}
    if status:
        query["status"] = status
    
    suggestions = await db.suggestions.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return suggestions

@app.get("/api/suggestions/count")
async def get_suggestions_count():
    """Get count of new suggestions"""
    total = await db.suggestions.count_documents({})
    new_count = await db.suggestions.count_documents({"status": "new"})
    return {"total": total, "new": new_count}

@app.put("/api/suggestions/{suggestion_id}/review")
async def review_suggestion(suggestion_id: str, status: str, reviewed_by: str = "Admin", review_notes: str = None):
    """Review a suggestion - set status to reviewed, implemented, or declined"""
    if status not in ["reviewed", "implemented", "declined"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.suggestions.update_one(
        {"id": suggestion_id},
        {"$set": {
            "status": status,
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
            "reviewed_by": reviewed_by,
            "review_notes": review_notes
        }}
    )
    await invalidate_cache()
    if result.modified_count > 0:
        return {"success": True, "message": f"Suggestion marked as {status}"}
    raise HTTPException(status_code=404, detail="Suggestion not found")

# Add comment to near miss
@app.post("/api/near-misses/{near_miss_id}/comment")
async def add_near_miss_comment(near_miss_id: str, comment: str, commented_by: str = "Admin"):
    """Add a comment to a near miss"""
    result = await db.near_misses.update_one(
        {"id": near_miss_id},
        {"$push": {"comments": {
            "text": comment,
            "by": commented_by,
            "at": datetime.now(timezone.utc).isoformat()
        }}}
    )
    if result.modified_count > 0:
        return {"success": True, "message": "Comment added"}
    raise HTTPException(status_code=404, detail="Near miss not found")

@app.put("/api/near-misses/{near_miss_id}/investigate")
async def investigate_near_miss(
    near_miss_id: str,
    severity: str = None,
    action_required: str = None,
    progress: str = None,
    investigation_notes: str = None,
    no_swp_or_not_covered: bool = False,
    swp_training_not_received: bool = False,
    trained_but_not_following: bool = False,
    investigated_by: str = None
):
    """Update investigation details for a near miss"""
    near_miss = await db.near_misses.find_one({"id": near_miss_id})
    if not near_miss:
        raise HTTPException(status_code=404, detail="Near miss not found")
    
    update_data = {
        "severity": severity,
        "action_required": action_required,
        "progress": progress,
        "investigation_notes": investigation_notes,
        "no_swp_or_not_covered": no_swp_or_not_covered,
        "swp_training_not_received": swp_training_not_received,
        "trained_but_not_following": trained_but_not_following,
        "investigated_by": investigated_by,
        "investigated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.near_misses.update_one({"id": near_miss_id}, {"$set": update_data})
    return {"success": True, "message": "Investigation updated"}

# Add comment to suggestion
@app.post("/api/suggestions/{suggestion_id}/comment")
async def add_suggestion_comment(suggestion_id: str, comment: str, commented_by: str = "Admin"):
    """Add a comment to a suggestion"""
    result = await db.suggestions.update_one(
        {"id": suggestion_id},
        {"$push": {"comments": {
            "text": comment,
            "by": commented_by,
            "at": datetime.now(timezone.utc).isoformat()
        }}}
    )
    if result.modified_count > 0:
        return {"success": True, "message": "Comment added"}
    raise HTTPException(status_code=404, detail="Suggestion not found")

# Near miss stats by location for pie chart
@app.get("/api/near-misses/stats/by-location")
async def get_near_misses_by_location():
    """Get near misses grouped by location for the last 4 months"""
    four_months_ago = (datetime.now(timezone.utc) - timedelta(days=120)).isoformat()
    
    pipeline = [
        {"$match": {"created_at": {"$gte": four_months_ago}}},
        {"$group": {
            "_id": {"$ifNull": ["$location", "Unknown"]},
            "count": {"$sum": 1}
        }},
        {"$project": {
            "location": "$_id",
            "count": 1,
            "_id": 0
        }}
    ]
    
    results = await db.near_misses.aggregate(pipeline).to_list(length=100)
    return results

# ============================================
# Accident Endpoints (Matching official accident record book)
# ============================================

@app.post("/api/accidents")
async def create_accident(accident: AccidentCreate, employee_number: str = None):
    """Report a new accident - matching official accident record book format"""
    # Generate report number
    count = await db.accidents.count_documents({})
    report_number = f"AR-{count + 1:04d}"
    
    accident_doc = {
        "id": str(uuid.uuid4()),
        "report_number": report_number,
        
        # Section 1: About the person who had the accident
        "injured_name": accident.injured_name,
        "injured_address": accident.injured_address,
        "injured_postcode": accident.injured_postcode,
        "injured_occupation": accident.injured_occupation,
        
        # Section 2: About you, the person filling in this record
        "reporter_name": accident.reporter_name,
        "reporter_address": accident.reporter_address,
        "reporter_postcode": accident.reporter_postcode,
        "reporter_occupation": accident.reporter_occupation,
        
        # Section 3: About the accident
        "accident_date": accident.accident_date,
        "accident_time": accident.accident_time,
        "accident_location": accident.accident_location,
        "accident_description": accident.accident_description,
        "injury_details": accident.injury_details,
        
        # Section 4: Employee consent
        "employee_consent": accident.employee_consent,
        
        # Section 5: RIDDOR (to be filled by employer)
        "riddor_reportable": False,
        "riddor_how_reported": None,
        "riddor_date_reported": None,
        
        # Additional
        "photos": accident.photos,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "new",
        "comments": []
    }
    await db.accidents.insert_one(accident_doc)
    await invalidate_cache()
    return {"success": True, "message": "Accident reported successfully", "id": accident_doc["id"], "report_number": report_number}

@app.get("/api/accidents")
async def get_accidents(status: str = None, limit: int = 100):
    """Get accident reports"""
    query = {}
    if status:
        query["status"] = status
    
    accidents = await db.accidents.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return accidents

@app.get("/api/accidents/count")
async def get_accidents_count():
    """Get count of accidents"""
    total = await db.accidents.count_documents({})
    new_count = await db.accidents.count_documents({"status": "new"})
    return {"total": total, "new": new_count}

@app.put("/api/accidents/{accident_id}/investigate")
async def investigate_accident(accident_id: str, status: str, investigated_by: str = "Admin", investigation_notes: str = None):
    """Update accident investigation status"""
    if status not in ["investigating", "closed"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.accidents.update_one(
        {"id": accident_id},
        {"$set": {
            "status": status,
            "investigated_at": datetime.now(timezone.utc).isoformat(),
            "investigated_by": investigated_by,
            "investigation_notes": investigation_notes
        }}
    )
    await invalidate_cache()
    if result.modified_count > 0:
        return {"success": True, "message": f"Accident marked as {status}"}
    raise HTTPException(status_code=404, detail="Accident not found")

@app.put("/api/accidents/{accident_id}/riddor")
async def update_riddor(accident_id: str, riddor_reportable: bool, how_reported: str = None, date_reported: str = None):
    """Update RIDDOR reporting details (employer only)"""
    result = await db.accidents.update_one(
        {"id": accident_id},
        {"$set": {
            "riddor_reportable": riddor_reportable,
            "riddor_how_reported": how_reported,
            "riddor_date_reported": date_reported
        }}
    )
    if result.modified_count > 0:
        return {"success": True, "message": "RIDDOR details updated"}
    raise HTTPException(status_code=404, detail="Accident not found")

@app.post("/api/accidents/{accident_id}/comment")
async def add_accident_comment(accident_id: str, comment: str, commented_by: str = "Admin"):
    """Add a comment to an accident"""
    result = await db.accidents.update_one(
        {"id": accident_id},
        {"$push": {"comments": {
            "text": comment,
            "by": commented_by,
            "at": datetime.now(timezone.utc).isoformat()
        }}}
    )
    if result.modified_count > 0:
        return {"success": True, "message": "Comment added"}
    raise HTTPException(status_code=404, detail="Accident not found")

# ============================================
# Whistleblowing Endpoints
# ============================================

@app.post("/api/whistleblowing")
async def create_whistleblow(whistleblow: WhistleblowCreate, employee_number: str = None):
    """Submit a whistleblowing report"""
    whistleblow_doc = {
        "id": str(uuid.uuid4()),
        "title": whistleblow.title,
        "description": whistleblow.description,
        "category": whistleblow.category,
        "location": whistleblow.location,
        "is_anonymous": whistleblow.is_anonymous,
        "submitted_by": whistleblow.submitted_by if not whistleblow.is_anonymous else None,
        "employee_number": employee_number if not whistleblow.is_anonymous else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "new",
        "comments": []
    }
    await db.whistleblowing.insert_one(whistleblow_doc)
    await invalidate_cache()
    return {"success": True, "message": "Report submitted successfully", "id": whistleblow_doc["id"]}

@app.get("/api/whistleblowing")
async def get_whistleblowing(status: str = None, limit: int = 100):
    """Get whistleblowing reports"""
    query = {}
    if status:
        query["status"] = status
    
    reports = await db.whistleblowing.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return reports

@app.get("/api/whistleblowing/count")
async def get_whistleblowing_count():
    """Get count of whistleblowing reports"""
    total = await db.whistleblowing.count_documents({})
    new_count = await db.whistleblowing.count_documents({"status": "new"})
    return {"total": total, "new": new_count}

@app.put("/api/whistleblowing/{report_id}/investigate")
async def investigate_whistleblow(report_id: str, status: str, investigated_by: str = "Admin", investigation_notes: str = None):
    """Update whistleblowing investigation status"""
    if status not in ["investigating", "resolved", "dismissed"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.whistleblowing.update_one(
        {"id": report_id},
        {"$set": {
            "status": status,
            "investigated_at": datetime.now(timezone.utc).isoformat(),
            "investigated_by": investigated_by,
            "investigation_notes": investigation_notes
        }}
    )
    await invalidate_cache()
    if result.modified_count > 0:
        return {"success": True, "message": f"Report marked as {status}"}
    raise HTTPException(status_code=404, detail="Report not found")

@app.post("/api/whistleblowing/{report_id}/comment")
async def add_whistleblow_comment(report_id: str, comment: str, commented_by: str = "Admin"):
    """Add a comment to a whistleblowing report"""
    result = await db.whistleblowing.update_one(
        {"id": report_id},
        {"$push": {"comments": {
            "text": comment,
            "by": commented_by,
            "at": datetime.now(timezone.utc).isoformat()
        }}}
    )
    if result.modified_count > 0:
        return {"success": True, "message": "Comment added"}
    raise HTTPException(status_code=404, detail="Report not found")

# ============================================
# Training Records Endpoints
# ============================================

@app.post("/api/training")
async def create_training_record(training: TrainingRecordCreate):
    """Create a new training record"""
    trainees_with_signatures = []
    for trainee in training.trainees:
        trainees_with_signatures.append({
            "employee_id": trainee.get("employee_id"),
            "employee_name": trainee.get("employee_name"),
            "is_agency": trainee.get("is_agency", False),
            "signed": False,
            "signed_at": None,
            "signature_data": None
        })
    
    training_doc = {
        "id": str(uuid.uuid4()),
        "swp_number": training.swp_number,
        "swp_version": training.swp_version,
        "department": training.department,
        "training_date": training.training_date,
        "notes": training.notes,
        "trainer_name": training.trainer_name,
        "trainer_employee_number": training.trainer_employee_number,
        "trainees": trainees_with_signatures,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending_signatures"
    }
    
    await db.training_records.insert_one(training_doc)
    return {"success": True, "message": "Training record created", "id": training_doc["id"]}

@app.get("/api/training")
async def get_training_records(status: str = None, limit: int = 100):
    """Get all training records"""
    query = {}
    if status:
        query["status"] = status
    records = await db.training_records.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return records

@app.get("/api/training/{record_id}")
async def get_training_record(record_id: str):
    """Get a specific training record"""
    record = await db.training_records.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Training record not found")
    return record

@app.get("/api/training/pending/{employee_number}")
async def get_pending_signatures(employee_number: str):
    """Get training records pending signature for a specific employee"""
    # Find records where this employee hasn't signed yet
    records = await db.training_records.find({
        "trainees": {
            "$elemMatch": {
                "employee_id": employee_number,
                "signed": False
            }
        }
    }, {"_id": 0}).to_list(length=100)
    return records

@app.put("/api/training/{record_id}/sign")
async def sign_training_record(record_id: str, employee_id: str = None, employee_name: str = None, signature_data: str = None):
    """Sign a training record"""
    record = await db.training_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Training record not found")
    
    # Find and update the trainee's signature
    trainees = record.get("trainees", [])
    updated = False
    all_signed = True
    
    for trainee in trainees:
        # Match by employee_id or employee_name for agency staff
        if (employee_id and trainee.get("employee_id") == employee_id) or \
           (employee_name and trainee.get("employee_name") == employee_name and trainee.get("is_agency")):
            trainee["signed"] = True
            trainee["signed_at"] = datetime.now(timezone.utc).isoformat()
            trainee["signature_data"] = signature_data
            updated = True
        
        if not trainee.get("signed"):
            all_signed = False
    
    if not updated:
        raise HTTPException(status_code=404, detail="Trainee not found in this record")
    
    # Update the record
    new_status = "completed" if all_signed else "pending_signatures"
    await db.training_records.update_one(
        {"id": record_id},
        {"$set": {"trainees": trainees, "status": new_status}}
    )
    
    return {"success": True, "message": "Signature recorded", "all_signed": all_signed}

@app.delete("/api/training/{record_id}")
async def delete_training_record(record_id: str):
    """Delete a training record"""
    result = await db.training_records.delete_one({"id": record_id})
    if result.deleted_count > 0:
        return {"success": True, "message": "Training record deleted"}
    raise HTTPException(status_code=404, detail="Training record not found")

@app.put("/api/training/{record_id}/sage-hr")
async def update_sage_hr_status(record_id: str, added: bool = True, updated_by: str = None):
    """Update the Sage HR status for a training record"""
    record = await db.training_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Training record not found")
    
    update_data = {
        "added_to_sage_hr": added,
        "added_to_sage_hr_at": datetime.now(timezone.utc).isoformat() if added else None,
        "added_to_sage_hr_by": updated_by if added else None
    }
    
    await db.training_records.update_one({"id": record_id}, {"$set": update_data})
    return {"success": True, "message": "Sage HR status updated"}

@app.get("/api/training/stats/count")
async def get_training_stats():
    """Get training record counts"""
    total = await db.training_records.count_documents({})
    pending = await db.training_records.count_documents({"status": "pending_signatures"})
    completed = await db.training_records.count_documents({"status": "completed"})
    return {"total": total, "pending": pending, "completed": completed}

# ============================================
# Work Progress Tracking Endpoints
# ============================================

@app.get("/api/jobs")
async def get_all_jobs():
    """Get all jobs with calculated stats for dashboard"""
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(length=1000)
    
    result = []
    for job in jobs:
        # Get all work entries for this job
        entries = await db.work_entries.find(
            {"job_id": job["id"]}, 
            {"_id": 0}
        ).sort("date_completed", 1).to_list(length=10000)
        
        # Calculate total completed
        total_completed = sum(e.get("hectares_completed", 0) for e in entries)
        area_left = max(0, job.get("total_area", 0) - total_completed)
        
        # Calculate Ha/day (average of daily entries)
        ha_per_day = 0
        if entries:
            # Group entries by date and calculate daily totals
            daily_totals = {}
            for entry in entries:
                date = entry.get("date_completed", "")[:10]  # Get just the date part
                if date:
                    daily_totals[date] = daily_totals.get(date, 0) + entry.get("hectares_completed", 0)
            
            if daily_totals:
                ha_per_day = sum(daily_totals.values()) / len(daily_totals)
        
        # Auto-update status to complete if area_left is 0
        if area_left <= 0 and job.get("status") == "active":
            await db.jobs.update_one(
                {"id": job["id"]},
                {"$set": {"status": "complete"}}
            )
            job["status"] = "complete"
        
        result.append({
            **job,
            "total_completed": round(total_completed, 2),
            "area_left": round(area_left, 2),
            "ha_per_day": round(ha_per_day, 2),
            "entries_count": len(entries),
            "last_entry": entries[-1] if entries else None
        })
    
    # Sort: active jobs first, then by name
    result.sort(key=lambda x: (0 if x["status"] == "active" else 1, x["name"]))
    
    return result

@app.post("/api/admin/jobs")
async def create_job(job_data: JobCreate):
    """Create a new job"""
    job = Job(
        name=job_data.name,
        total_area=job_data.total_area,
        target_date=job_data.target_date
    )
    
    await db.jobs.insert_one(job.dict())
    
    return {
        "success": True,
        "message": f"Job '{job.name}' created successfully",
        "job": job.dict()
    }

@app.get("/api/admin/jobs/{job_id}")
async def get_job_details(job_id: str):
    """Get detailed job info including all work entries"""
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    entries = await db.work_entries.find(
        {"job_id": job_id}, 
        {"_id": 0}
    ).sort("date_completed", -1).to_list(length=10000)
    
    total_completed = sum(e.get("hectares_completed", 0) for e in entries)
    
    return {
        **job,
        "total_completed": round(total_completed, 2),
        "area_left": round(max(0, job.get("total_area", 0) - total_completed), 2),
        "entries": entries
    }

@app.post("/api/admin/jobs/{job_id}/work-entry")
async def add_work_entry(job_id: str, entry_data: WorkEntryCreate):
    """Add a work entry to a job"""
    # Verify job exists
    job = await db.jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Use provided date or today
    date_completed = entry_data.date_completed or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    entry = WorkEntry(
        job_id=job_id,
        hectares_completed=entry_data.hectares_completed,
        date_completed=date_completed,
        entered_by=entry_data.entered_by
    )
    
    await db.work_entries.insert_one(entry.dict())
    
    # Check if job should be marked complete
    entries = await db.work_entries.find({"job_id": job_id}, {"_id": 0}).to_list(length=10000)
    total_completed = sum(e.get("hectares_completed", 0) for e in entries)
    area_left = max(0, job.get("total_area", 0) - total_completed)
    
    if area_left <= 0:
        await db.jobs.update_one({"id": job_id}, {"$set": {"status": "complete"}})
    
    return {
        "success": True,
        "message": f"Added {entry_data.hectares_completed} Ha to '{job['name']}'",
        "entry": entry.dict(),
        "total_completed": round(total_completed, 2),
        "area_left": round(area_left, 2)
    }

@app.delete("/api/admin/jobs/{job_id}")
async def delete_job(job_id: str):
    """Delete a job and all its work entries"""
    job = await db.jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Delete job and all related entries
    await db.jobs.delete_one({"id": job_id})
    deleted_entries = await db.work_entries.delete_many({"job_id": job_id})
    
    return {
        "success": True,
        "message": f"Job '{job['name']}' deleted",
        "entries_deleted": deleted_entries.deleted_count
    }

@app.put("/api/admin/jobs/{job_id}")
async def update_job(job_id: str, job_data: JobCreate):
    """Update a job's name, total area, or target date"""
    job = await db.jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {"name": job_data.name, "total_area": job_data.total_area, "target_date": job_data.target_date}}
    )
    
    return {
        "success": True,
        "message": f"Job updated successfully"
    }

@app.put("/api/admin/jobs/{job_id}/reopen")
async def reopen_job(job_id: str):
    """Reopen a completed job (set status back to active)"""
    job = await db.jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    await db.jobs.update_one({"id": job_id}, {"$set": {"status": "active"}})
    
    return {
        "success": True,
        "message": f"Job '{job['name']}' reopened"
    }

@app.delete("/api/admin/work-entries/{entry_id}")
async def delete_work_entry(entry_id: str):
    """Delete a specific work entry"""
    entry = await db.work_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    await db.work_entries.delete_one({"id": entry_id})
    
    # Check if parent job should be reopened
    job = await db.jobs.find_one({"id": entry["job_id"]})
    if job and job.get("status") == "complete":
        entries = await db.work_entries.find({"job_id": job["id"]}, {"_id": 0}).to_list(length=10000)
        total_completed = sum(e.get("hectares_completed", 0) for e in entries)
        if total_completed < job.get("total_area", 0):
            await db.jobs.update_one({"id": job["id"]}, {"$set": {"status": "active"}})
    
    return {
        "success": True,
        "message": "Work entry deleted"
    }

# ============ EXCEL EXPORT ENDPOINTS FOR NEW FEATURES ============

@app.get("/api/near-misses/export/excel")
async def export_near_misses_excel():
    """Export all near misses to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    near_misses = await db.near_misses.find({}, {"_id": 0}).to_list(length=10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Near Misses"
    
    # Headers
    headers = ["ID", "Date", "Location", "Description", "Submitted By", "Anonymous", "Acknowledged", "Acknowledged By", "Acknowledged Date"]
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_num, nm in enumerate(near_misses, 2):
        ws.cell(row=row_num, column=1, value=nm.get("id", ""))
        ws.cell(row=row_num, column=2, value=nm.get("created_at", "")[:10] if nm.get("created_at") else "")
        ws.cell(row=row_num, column=3, value=nm.get("location", ""))
        ws.cell(row=row_num, column=4, value=nm.get("description", ""))
        ws.cell(row=row_num, column=5, value=nm.get("submitted_by", "") if not nm.get("is_anonymous") else "Anonymous")
        ws.cell(row=row_num, column=6, value="Yes" if nm.get("is_anonymous") else "No")
        ws.cell(row=row_num, column=7, value="Yes" if nm.get("acknowledged") else "No")
        ws.cell(row=row_num, column=8, value=nm.get("acknowledged_by", ""))
        ws.cell(row=row_num, column=9, value=nm.get("acknowledged_at", "")[:10] if nm.get("acknowledged_at") else "")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=near_misses.xlsx"}
    )

@app.get("/api/suggestions/export/excel")
async def export_suggestions_excel():
    """Export all suggestions to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    suggestions = await db.suggestions.find({}, {"_id": 0}).to_list(length=10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Suggestions"
    
    # Headers
    headers = ["ID", "Date", "Title", "Category", "Location", "Description", "Submitted By", "Anonymous", "Status", "Reviewed By", "Review Notes"]
    header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_num, sg in enumerate(suggestions, 2):
        ws.cell(row=row_num, column=1, value=sg.get("id", ""))
        ws.cell(row=row_num, column=2, value=sg.get("created_at", "")[:10] if sg.get("created_at") else "")
        ws.cell(row=row_num, column=3, value=sg.get("title", ""))
        ws.cell(row=row_num, column=4, value=sg.get("category", ""))
        ws.cell(row=row_num, column=5, value=sg.get("location", ""))
        ws.cell(row=row_num, column=6, value=sg.get("description", ""))
        ws.cell(row=row_num, column=7, value=sg.get("submitted_by", "") if not sg.get("is_anonymous") else "Anonymous")
        ws.cell(row=row_num, column=8, value="Yes" if sg.get("is_anonymous") else "No")
        ws.cell(row=row_num, column=9, value=sg.get("status", "new").capitalize())
        ws.cell(row=row_num, column=10, value=sg.get("reviewed_by", ""))
        ws.cell(row=row_num, column=11, value=sg.get("review_notes", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 40
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=suggestions.xlsx"}
    )

@app.get("/api/accidents/export/excel")
async def export_accidents_excel():
    """Export all accidents to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    accidents = await db.accidents.find({}, {"_id": 0}).to_list(length=10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Accidents"
    
    # Headers
    headers = [
        "Report No", "Date", "Time", "Location", 
        "Injured Name", "Injured Occupation", "Injured Address",
        "Reporter Name", "Reporter Occupation",
        "Accident Description", "Injury Details",
        "Employee Consent", "RIDDOR Reportable", "RIDDOR How Reported", "RIDDOR Date Reported",
        "Status", "Investigation Notes", "Investigated By"
    ]
    header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_num, acc in enumerate(accidents, 2):
        ws.cell(row=row_num, column=1, value=acc.get("report_number", ""))
        ws.cell(row=row_num, column=2, value=acc.get("accident_date", ""))
        ws.cell(row=row_num, column=3, value=acc.get("accident_time", ""))
        ws.cell(row=row_num, column=4, value=acc.get("accident_location", ""))
        ws.cell(row=row_num, column=5, value=acc.get("injured_name", ""))
        ws.cell(row=row_num, column=6, value=acc.get("injured_occupation", ""))
        ws.cell(row=row_num, column=7, value=f"{acc.get('injured_address', '')} {acc.get('injured_postcode', '')}".strip())
        ws.cell(row=row_num, column=8, value=acc.get("reporter_name", ""))
        ws.cell(row=row_num, column=9, value=acc.get("reporter_occupation", ""))
        ws.cell(row=row_num, column=10, value=acc.get("accident_description", ""))
        ws.cell(row=row_num, column=11, value=acc.get("injury_details", ""))
        ws.cell(row=row_num, column=12, value="Yes" if acc.get("employee_consent") else "No")
        ws.cell(row=row_num, column=13, value="Yes" if acc.get("riddor_reportable") else "No")
        ws.cell(row=row_num, column=14, value=acc.get("riddor_how_reported", ""))
        ws.cell(row=row_num, column=15, value=acc.get("riddor_date_reported", ""))
        ws.cell(row=row_num, column=16, value=acc.get("status", "new").capitalize())
        ws.cell(row=row_num, column=17, value=acc.get("investigation_notes", ""))
        ws.cell(row=row_num, column=18, value=acc.get("investigated_by", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 16
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 18
    ws.column_dimensions['P'].width = 14
    ws.column_dimensions['Q'].width = 40
    ws.column_dimensions['R'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=accidents.xlsx"}
    )

@app.get("/api/whistleblowing/export/excel")
async def export_whistleblowing_excel():
    """Export all whistleblowing reports to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    reports = await db.whistleblowing.find({}, {"_id": 0}).to_list(length=10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Whistleblowing"
    
    # Headers
    headers = ["ID", "Date", "Title", "Category", "Location", "Description", "Submitted By", "Anonymous", "Status", "Investigation Notes", "Investigated By"]
    header_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_num, rp in enumerate(reports, 2):
        ws.cell(row=row_num, column=1, value=rp.get("id", ""))
        ws.cell(row=row_num, column=2, value=rp.get("created_at", "")[:10] if rp.get("created_at") else "")
        ws.cell(row=row_num, column=3, value=rp.get("title", ""))
        ws.cell(row=row_num, column=4, value=rp.get("category", ""))
        ws.cell(row=row_num, column=5, value=rp.get("location", ""))
        ws.cell(row=row_num, column=6, value=rp.get("description", ""))
        ws.cell(row=row_num, column=7, value=rp.get("submitted_by", "") if not rp.get("is_anonymous") else "Anonymous")
        ws.cell(row=row_num, column=8, value="Yes" if rp.get("is_anonymous") else "No")
        ws.cell(row=row_num, column=9, value=rp.get("status", "new").capitalize())
        ws.cell(row=row_num, column=10, value=rp.get("investigation_notes", ""))
        ws.cell(row=row_num, column=11, value=rp.get("investigated_by", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=whistleblowing.xlsx"}
    )

# DAILY WORKPLAN ENDPOINTS
# Collections:
#   db.workplan        -> single doc {key:'current', week_start, draft_rows, published_rows, published_week_start, published_at}
#   db.workplan_jobs   -> {id, name, order}
#   db.workplan_colors -> {id, name, color, order}
# ==========================================================================

class WorkplanSaveRequest(BaseModel):
    week_start: str
    rows: List[dict] = []

class JobItem(BaseModel):
    name: str

class ColorItem(BaseModel):
    name: str
    color: str

@app.get("/api/workplan")
async def get_workplan():
    """Return the draft workplan that managers edit."""
    doc = await db.workplan.find_one({"key": "current"}, {"_id": 0})
    if not doc:
        return {"week_start": None, "rows": [], "published_at": None, "is_published": False}
    return {
        "week_start": doc.get("week_start"),
        "rows": doc.get("draft_rows", []),
        "published_at": doc.get("published_at"),
        "is_published": bool(doc.get("published_rows"))
    }

@app.put("/api/workplan")
async def save_workplan(req: WorkplanSaveRequest):
    """Save the draft workplan. Auto-archives previous week if week_start changed."""
    # Check if week changed — archive the old one
    existing = await db.workplan.find_one({"key": "current"}, {"_id": 0})
    if existing and existing.get("week_start") and existing.get("week_start") != req.week_start:
        old_rows = existing.get("draft_rows", [])
        if old_rows:
            await db.workplan_archive.update_one(
                {"week_start": existing["week_start"]},
                {"$set": {
                    "week_start": existing["week_start"],
                    "rows": old_rows,
                    "archived_at": datetime.now(timezone.utc).isoformat()
                }},
                upsert=True
            )
    await db.workplan.update_one(
        {"key": "current"},
        {"$set": {
            "key": "current",
            "week_start": req.week_start,
            "draft_rows": req.rows,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    return {"success": True}

@app.post("/api/workplan/publish")
async def publish_workplan():
    """Publish the current draft so staff see it on the dashboard."""
    doc = await db.workplan.find_one({"key": "current"})
    if not doc:
        raise HTTPException(status_code=404, detail="No workplan to publish")
    now = datetime.now(timezone.utc).isoformat()
    await db.workplan.update_one(
        {"key": "current"},
        {"$set": {
            "published_rows": doc.get("draft_rows", []),
            "published_week_start": doc.get("week_start"),
            "published_at": now
        }}
    )
    return {"success": True, "published_at": now}

# Workplan presence tracking - in-memory for simplicity (resets on server restart)
workplan_active_users = {}  # {user_id: {name: str, last_seen: datetime}}
PRESENCE_TIMEOUT_SECONDS = 60  # Consider user gone after 60 seconds

class PresenceRequest(BaseModel):
    user_id: str
    user_name: str

@app.post("/api/workplan/presence/heartbeat")
async def workplan_presence_heartbeat(req: PresenceRequest):
    """Update user's presence in the workplan editor."""
    now = datetime.now(timezone.utc)
    workplan_active_users[req.user_id] = {
        "name": req.user_name,
        "last_seen": now
    }
    
    # Clean up stale users
    stale_cutoff = now - timedelta(seconds=PRESENCE_TIMEOUT_SECONDS)
    stale_users = [uid for uid, data in workplan_active_users.items() 
                   if data["last_seen"] < stale_cutoff]
    for uid in stale_users:
        del workplan_active_users[uid]
    
    # Return list of other active users
    other_users = [
        {"user_id": uid, "name": data["name"]}
        for uid, data in workplan_active_users.items()
        if uid != req.user_id
    ]
    
    return {"active_users": other_users}

@app.post("/api/workplan/presence/leave")
async def workplan_presence_leave(req: PresenceRequest):
    """Remove user from active users when they leave the page."""
    if req.user_id in workplan_active_users:
        del workplan_active_users[req.user_id]
    return {"success": True}

@app.get("/api/workplan/published")
async def get_published_workplan():
    """Return the published workplan for the staff dashboard view."""
    doc = await db.workplan.find_one({"key": "current"}, {"_id": 0})
    if not doc or not doc.get("published_rows"):
        return {"week_start": None, "rows": [], "published_at": None}
    return {
        "week_start": doc.get("published_week_start"),
        "rows": doc.get("published_rows", []),
        "published_at": doc.get("published_at")
    }

@app.get("/api/workplan/jobs")
async def get_workplan_jobs():
    jobs = await db.workplan_jobs.find({}, {"_id": 0}).sort("order", 1).to_list(length=None)
    return jobs

@app.post("/api/workplan/jobs")
async def add_workplan_job(item: JobItem):
    if not item.name.strip():
        raise HTTPException(status_code=400, detail="Job name is required")
    count = await db.workplan_jobs.count_documents({})
    job = {"id": str(uuid.uuid4()), "name": item.name.strip(), "order": count}
    await db.workplan_jobs.insert_one({**job})
    return job

@app.delete("/api/workplan/jobs/{job_id}")
async def delete_workplan_job(job_id: str):
    await db.workplan_jobs.delete_one({"id": job_id})
    return {"success": True}

@app.get("/api/workplan/colors")
async def get_workplan_colors():
    colors = await db.workplan_colors.find({}, {"_id": 0}).sort("order", 1).to_list(length=None)
    return colors

@app.post("/api/workplan/colors")
async def add_workplan_color(item: ColorItem):
    if not item.name.strip():
        raise HTTPException(status_code=400, detail="Colour name is required")
    count = await db.workplan_colors.count_documents({})
    c = {"id": str(uuid.uuid4()), "name": item.name.strip(), "color": item.color, "order": count}
    await db.workplan_colors.insert_one({**c})
    return c

@app.put("/api/workplan/colors/{color_id}")
async def update_workplan_color(color_id: str, item: ColorItem):
    await db.workplan_colors.update_one(
        {"id": color_id},
        {"$set": {"name": item.name.strip(), "color": item.color}}
    )
    return {"success": True}

@app.delete("/api/workplan/colors/{color_id}")
async def delete_workplan_color(color_id: str):
    await db.workplan_colors.delete_one({"id": color_id})
    return {"success": True}

@app.post("/api/admin/workplan/import-staff")
async def import_workplan_staff():
    """Import staff rows from the uploaded original Excel workplan into the current week's workplan.
    Parses daily assignments, marks leavers, fuzzy-matches job names."""
    import openpyxl, re
    from io import BytesIO
    from difflib import get_close_matches
    
    excel_url = "https://customer-assets.emergentagent.com/job_3e1cee5c-63e2-4d27-9a1e-16878b2e56b8/artifacts/ls1wpmcs_Daily%2520Workplan%25202024%20%28version%201%29.xlsb.xlsx"
    
    import httpx
    async with httpx.AsyncClient() as client:
        resp = await client.get(excel_url)
        if resp.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to download Excel file")
        wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
    
    # Build job name lookup for fuzzy matching
    all_jobs_docs = await db.workplan_jobs.find({}, {"_id": 0}).to_list(length=None)
    job_names = [j["name"] for j in all_jobs_docs]
    job_names_lower = {j.lower().strip(): j for j in job_names}
    
    # Get colour categories for auto-assignment
    all_colors = await db.workplan_colors.find({}, {"_id": 0}).to_list(length=None)
    color_by_name = {c["name"].lower(): c for c in all_colors}
    
    def fuzzy_match_job(raw):
        if not raw:
            return ''
        raw_clean = raw.strip()
        if not raw_clean:
            return ''
        # Exact match (case-insensitive)
        if raw_clean.lower() in job_names_lower:
            return job_names_lower[raw_clean.lower()]
        # Close match
        matches = get_close_matches(raw_clean.lower(), list(job_names_lower.keys()), n=1, cutoff=0.6)
        if matches:
            return job_names_lower[matches[0]]
        # Partial match
        for jn_lower, jn in job_names_lower.items():
            if raw_clean.lower() in jn_lower or jn_lower in raw_clean.lower():
                return jn
        return raw_clean  # Return as-is if no match
    
    def auto_color_for_notes(notes_text):
        """Try to assign a colour based on field/notes text."""
        if not notes_text:
            return '', ''
        nl = notes_text.lower()
        for cname, cdata in color_by_name.items():
            # Check if crop/area name appears in notes
            if cname.replace('/', '').replace(' ', '') in nl.replace('/', '').replace(' ', ''):
                return cdata.get('color', ''), cdata.get('id', '')
        # Common keywords
        if 'larkshall' in nl:
            c = color_by_name.get('larkshall')
            if c: return c['color'], c['id']
        if 'snetterton' in nl:
            c = color_by_name.get('snetterton')
            if c: return c['color'], c['id']
        if 'onion' in nl:
            c = color_by_name.get('onions')
            if c: return c['color'], c['id']
        if 'carrot' in nl:
            c = color_by_name.get('carrots')
            if c: return c['color'], c['id']
        if 'potato' in nl or 'spud' in nl:
            c = color_by_name.get('potatoes')
            if c: return c['color'], c['id']
        return '', ''
    
    rows = []
    jcb_reached = False
    
    # Parse Main Sheet
    ws_main = wb['Main Sheet']
    for r in range(3, ws_main.max_row + 1):
        vehicle = ws_main.cell(r, 1).value
        employee = ws_main.cell(r, 2).value
        manager = ws_main.cell(r, 3).value
        start_time = ws_main.cell(r, 4).value
        field_notes = ws_main.cell(r, 6).value
        
        if not employee and not vehicle:
            continue
        
        emp = str(employee).strip() if employee else ''
        veh = str(vehicle).strip() if vehicle else ''
        mgr = str(manager).strip() if manager else ''
        if mgr.startswith('zzzz'):
            mgr = ''
        
        st = ''
        if start_time:
            st_str = str(start_time)
            if ':' in st_str:
                parts = st_str.split(':')
                st = f"{parts[0]}:{parts[1]}"
        
        fn = str(field_notes).strip() if field_notes else ''
        
        # Detect JCB rows and mark everything below as left
        if emp.upper().startswith('JCB'):
            jcb_reached = True
        
        is_left = jcb_reached
        
        # Parse daily job assignments (cols 7-20: 7 days x 2 cols AM/PM)
        # Wed=7/8, Thu=9/10, Fri=11/12, Sat=13/14, Sun=15/16, Mon=17/18, Tue=19/20
        # Map to our 0-6 (Mon-Sun) index
        excel_day_cols = {
            0: (17, 18),  # Monday
            1: (19, 20),  # Tuesday
            2: (7, 8),    # Wednesday
            3: (9, 10),   # Thursday
            4: (11, 12),  # Friday
            5: (13, 14),  # Saturday
            6: (15, 16),  # Sunday
        }
        
        note_color, note_color_id = auto_color_for_notes(fn)
        
        days = {}
        for day_idx in range(7):
            am_col, pm_col = excel_day_cols[day_idx]
            am_raw = ws_main.cell(r, am_col).value
            pm_raw = ws_main.cell(r, pm_col).value
            
            am_job = fuzzy_match_job(str(am_raw) if am_raw else '')
            pm_job = fuzzy_match_job(str(pm_raw) if pm_raw else '')
            
            days[str(day_idx)] = {
                'am': {'job': am_job, 'color': note_color, 'color_id': note_color_id},
                'pm': {'job': pm_job, 'color': note_color, 'color_id': note_color_id}
            }
        
        rows.append({
            'id': str(uuid.uuid4()),
            'employee_name': emp,
            'vehicle': veh,
            'implement': '',
            'manager': mgr,
            'start_time': st,
            'notes': fn,
            'groupColor': '',
            'left': is_left,
            'days': days
        })
    
    # Parse Harvest Staff
    if 'Harvest Staff' in wb.sheetnames:
        ws_harvest = wb['Harvest Staff']
        # Harvest days start at col 6 (Sun), col 7 (Mon), ..., col 12 (Sat)
        harvest_day_cols = {
            0: 7,   # Monday
            1: 8,   # Tuesday
            2: 9,   # Wednesday
            3: 10,  # Thursday
            4: 11,  # Friday
            5: 12,  # Saturday
            6: 6,   # Sunday
        }
        
        for r in range(3, ws_harvest.max_row + 1):
            role = ws_harvest.cell(r, 2).value
            name = ws_harvest.cell(r, 3).value
            manager = ws_harvest.cell(r, 4).value
            at_workshop = ws_harvest.cell(r, 5).value
            
            if not name:
                continue
            
            emp = str(name).strip()
            mgr = str(manager).strip() if manager else ''
            role_str = str(role).strip() if role else ''
            at_ws = str(at_workshop).strip() if at_workshop else ''
            
            st = ''
            if at_ws:
                m = re.match(r'(\d+[:.]\d+\s*[AaPp][Mm]?)', at_ws)
                if m:
                    st = m.group(1).replace('.', ':')
            
            note_color, note_color_id = auto_color_for_notes(at_ws)
            
            days = {}
            for day_idx in range(7):
                col = harvest_day_cols[day_idx]
                raw = ws_harvest.cell(r, col).value
                job = fuzzy_match_job(str(raw) if raw else '')
                days[str(day_idx)] = {
                    'am': {'job': job, 'color': note_color, 'color_id': note_color_id},
                    'pm': {'job': job, 'color': note_color, 'color_id': note_color_id}
                }
            
            rows.append({
                'id': str(uuid.uuid4()),
                'employee_name': emp,
                'vehicle': role_str,
                'implement': '',
                'manager': mgr,
                'start_time': st,
                'notes': at_ws,
                'groupColor': '',
                'left': False,
                'days': days
            })
    
    # Update jobs from Excel
    if 'Jobs' in wb.sheetnames:
        ws_jobs = wb['Jobs']
        jobs_from_excel = []
        for r in range(1, ws_jobs.max_row + 1):
            v = ws_jobs.cell(r, 1).value
            if v and str(v).strip():
                jobs_from_excel.append(str(v).strip())
        if 'Wet Day Jobs' in wb.sheetnames:
            ws_wet = wb['Wet Day Jobs']
            for r in range(1, ws_wet.max_row + 1):
                v = ws_wet.cell(r, 1).value
                if v and str(v).strip():
                    jobs_from_excel.append(str(v).strip())
        seen = set()
        unique_jobs = []
        for j in jobs_from_excel:
            if j not in seen:
                seen.add(j)
                unique_jobs.append(j)
        await db.workplan_jobs.delete_many({})
        for i, name in enumerate(unique_jobs):
            await db.workplan_jobs.insert_one({"id": str(uuid.uuid4()), "name": name, "order": i})
    
    # Save workplan
    from datetime import datetime, timezone
    today = datetime.now(timezone.utc)
    monday = today - timedelta(days=today.weekday())
    week_start = monday.strftime('%Y-%m-%d')
    
    await db.workplan.update_one(
        {"key": "current"},
        {"$set": {
            "key": "current",
            "week_start": week_start,
            "draft_rows": rows,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    active = sum(1 for r in rows if not r.get('left', False))
    left = sum(1 for r in rows if r.get('left', False))
    
    return {
        "success": True,
        "total_rows": len(rows),
        "active_rows": active,
        "left_rows": left,
        "jobs_updated": True
    }

@app.get("/api/workplan/costing")
async def get_workplan_costing(from_date: str = None, until_date: str = None):
    """Calculate % breakdown by (area/crop + job). Aggregates current + archived weeks.
    Optional from_date/until_date (YYYY-MM-DD) filter which days to include."""
    
    all_colors = await db.workplan_colors.find({}, {"_id": 0}).to_list(length=None)
    color_lookup = {c["id"]: c for c in all_colors}
    color_by_hex = {c["color"]: c for c in all_colors}
    
    # Collect all week documents (current + archives)
    week_docs = []
    current = await db.workplan.find_one({"key": "current"}, {"_id": 0})
    if current and current.get("draft_rows"):
        week_docs.append({"week_start": current.get("week_start"), "rows": current.get("draft_rows", [])})
    archives = await db.workplan_archive.find({}, {"_id": 0}).to_list(length=None)
    for a in archives:
        week_docs.append({"week_start": a.get("week_start"), "rows": a.get("rows", [])})
    
    combined = {}
    left_combined = {}
    total_cells = 0
    left_total = 0
    
    fd = None
    ud = None
    if from_date:
        try: fd = datetime.strptime(from_date, "%Y-%m-%d").date()
        except ValueError: pass
    if until_date:
        try: ud = datetime.strptime(until_date, "%Y-%m-%d").date()
        except ValueError: pass
    
    for week_doc in week_docs:
        ws = week_doc.get("week_start")
        if not ws:
            continue
        try:
            week_monday = datetime.strptime(ws, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        
        for row in week_doc.get("rows", []):
            is_left = row.get('left', False)
            days = row.get('days', {})
            if isinstance(days, list):
                day_items = list(enumerate(days))
            else:
                day_items = [(int(k), v) for k, v in days.items()]
            
            for day_idx, day_data in day_items:
                day_date = week_monday + timedelta(days=int(day_idx))
                if fd and day_date < fd:
                    continue
                if ud and day_date > ud:
                    continue
                
                for period in ['am', 'pm']:
                    cell = day_data.get(period, {})
                    job = cell.get('job', '').strip()
                    if not job:
                        continue
                    color = cell.get('color', '')
                    color_id = cell.get('color_id', '')
                    area = 'Unassigned'
                    if color_id and color_id in color_lookup:
                        area = color_lookup[color_id]['name']
                    elif color and color in color_by_hex:
                        area = color_by_hex[color]['name']
                    key = f"{area}, {job}"
                    if is_left:
                        left_total += 1
                        left_combined[key] = left_combined.get(key, 0) + 1
                    else:
                        total_cells += 1
                        combined[key] = combined.get(key, 0) + 1
    
    def to_breakdown(counts, total):
        return sorted([
            {"name": k, "area": k.split(", ", 1)[0], "job": k.split(", ", 1)[1] if ", " in k else k,
             "count": v, "percent": round(v / total * 100, 1) if total > 0 else 0}
            for k, v in counts.items()
        ], key=lambda x: -x["count"])
    
    return {
        "combined_breakdown": to_breakdown(combined, total_cells),
        "total_cells": total_cells,
        "left_combined_breakdown": to_breakdown(left_combined, left_total),
        "left_total_cells": left_total,
        "weeks_included": len(week_docs)
    }

# --- One-time data migration: import JSON exports from the old Emergent app ---
import json as _json
from fastapi.responses import HTMLResponse

IMPORTABLE_COLLECTIONS = ["assets", "checklist_templates", "staff", "repair_status", "checklists"]


def _admin_password_ok(password: str) -> bool:
    expected = os.environ.get("REACT_APP_ADMIN_PASSWORD")
    return bool(expected) and password == expected


@app.post("/api/admin/import-data")
async def import_data(
    collection: str = Form(...),
    password: str = Form(...),
    file: UploadFile = File(...),
):
    """Replace one collection's contents with records from an exported JSON file."""
    if not _admin_password_ok(password):
        raise HTTPException(status_code=401, detail="Wrong admin password.")
    if collection not in IMPORTABLE_COLLECTIONS:
        raise HTTPException(status_code=400, detail=f"Unknown data type '{collection}'.")

    raw = await file.read()
    try:
        docs = _json.loads(raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"That file is not valid JSON: {e}")
    if not isinstance(docs, list):
        raise HTTPException(status_code=400, detail="Expected a JSON file containing a list of records.")

    for d in docs:
        if isinstance(d, dict):
            d.pop("_id", None)  # let MongoDB assign fresh internal ids

    coll = db[collection]
    deleted = (await coll.delete_many({})).deleted_count

    inserted = 0
    batch = []
    for d in docs:
        batch.append(d)
        if len(batch) >= 200:
            await coll.insert_many(batch)
            inserted += len(batch)
            batch = []
    if batch:
        await coll.insert_many(batch)
        inserted += len(batch)

    try:
        await invalidate_cache()
    except Exception:
        pass

    return {"collection": collection, "replaced": deleted, "imported": inserted}


@app.get("/import", response_class=HTMLResponse)
async def import_page():
    """Simple browser page for importing exported data files."""
    return """<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Data Import - Machine Checklist</title>
<style>
  body { font-family: system-ui, sans-serif; max-width: 640px; margin: 40px auto; padding: 0 16px; color: #1a202c; }
  h1 { font-size: 1.4rem; } .card { border: 1px solid #cbd5e0; border-radius: 8px; padding: 16px; margin: 12px 0; }
  input, button { font-size: 1rem; padding: 8px; margin: 4px 0; }
  button { background: #2f855a; color: #fff; border: 0; border-radius: 6px; padding: 10px 18px; cursor: pointer; }
  button:disabled { background: #a0aec0; }
  .ok { color: #2f855a; } .err { color: #c53030; } li { margin: 4px 0; }
</style>
</head>
<body>
<h1>Import data from the old app</h1>
<p>Select the exported <b>.json</b> files (assets, checklists, staff, repair_status,
checklist_templates). Each file <b>replaces</b> that data type entirely, so only pick
the ones you mean to bring across.</p>
<div class="card">
  <label>Admin password<br><input type="password" id="pw" style="width:100%"></label><br>
  <label>Export files<br><input type="file" id="files" multiple accept=".json"></label><br>
  <button id="go">Import selected files</button>
</div>
<ul id="log"></ul>
<script>
document.getElementById('go').onclick = async function () {
  const pw = document.getElementById('pw').value;
  const files = document.getElementById('files').files;
  const log = document.getElementById('log');
  const btn = this;
  if (!pw) { alert('Enter the admin password first.'); return; }
  if (!files.length) { alert('Choose at least one .json export file.'); return; }
  btn.disabled = true;
  for (const f of files) {
    const name = f.name.replace(/\\.json$/i, '');
    const li = document.createElement('li');
    li.textContent = f.name + ' — importing...';
    log.appendChild(li);
    const fd = new FormData();
    fd.append('collection', name);
    fd.append('password', pw);
    fd.append('file', f);
    try {
      const r = await fetch('/api/admin/import-data', { method: 'POST', body: fd });
      const j = await r.json();
      if (r.ok) {
        li.className = 'ok';
        li.textContent = f.name + ' — done: ' + j.imported + ' records imported (replaced ' + j.replaced + ')';
      } else {
        li.className = 'err';
        li.textContent = f.name + ' — failed: ' + (j.detail || r.status);
      }
    } catch (e) {
      li.className = 'err';
      li.textContent = f.name + ' — failed: ' + e;
    }
  }
  btn.disabled = false;
};
</script>
</body>
</html>"""


# --- Serve the React frontend build (single-service deployment) ---
# When the frontend has been built (frontend/build exists), the backend serves it
# directly, so the whole app runs as ONE service. During local development with
# the React dev server this block is skipped automatically.
from pathlib import Path
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

FRONTEND_BUILD = Path(__file__).resolve().parent.parent / "frontend" / "build"
if FRONTEND_BUILD.is_dir():
    app.mount("/static", StaticFiles(directory=FRONTEND_BUILD / "static"), name="static")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        candidate = FRONTEND_BUILD / full_path
        if full_path and candidate.is_file():
            return FileResponse(candidate)
        return FileResponse(FRONTEND_BUILD / "index.html")



if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)