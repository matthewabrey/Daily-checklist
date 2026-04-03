"""
SharePoint Auto-Sync Service for Staff List
Uses client credentials flow (app-only authentication) for scheduled background sync.
"""

import os
import requests
import logging
from typing import List, Dict, Tuple
from io import BytesIO
import openpyxl
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger(__name__)

class SharePointAutoSync:
    def __init__(self):
        self.client_id = os.environ.get('AZURE_CLIENT_ID')
        self.client_secret = os.environ.get('AZURE_CLIENT_SECRET')
        self.tenant_id = os.environ.get('AZURE_TENANT_ID')
        self.site_url = os.environ.get('SHAREPOINT_SITE_URL', 'https://rgafarms.sharepoint.com/sites/Crops')
        self.staff_filename = os.environ.get('SHAREPOINT_STAFF_FILENAME', 'Name List.xlsx')
        self.assets_filename = os.environ.get('SHAREPOINT_ASSETS_FILENAME', 'AssetList.xlsx')
        self.folder_path = 'General/Apps/Checklist App'  # Folder path within the document library
        
        self.token_url = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        self.graph_url = "https://graph.microsoft.com/v1.0"
        self.access_token = None
        
    def _get_access_token(self) -> str:
        """Get access token using client credentials flow (app-only)"""
        if not self.client_id:
            raise ValueError("Missing AZURE_CLIENT_ID environment variable")
        if not self.client_secret:
            raise ValueError("Missing AZURE_CLIENT_SECRET environment variable")
        if not self.tenant_id:
            raise ValueError("Missing AZURE_TENANT_ID environment variable")
        
        logger.info(f"Attempting to get access token with client_id length: {len(self.client_id)}, tenant_id length: {len(self.tenant_id)}")
        
        data = {
            'grant_type': 'client_credentials',
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'scope': 'https://graph.microsoft.com/.default'
        }
        
        response = requests.post(self.token_url, data=data, timeout=30)
        
        if response.status_code != 200:
            logger.error(f"Token request failed: {response.status_code} - {response.text}")
            raise Exception(f"Failed to get access token: {response.text}")
        
        token_data = response.json()
        self.access_token = token_data['access_token']
        logger.info("Successfully acquired access token via client credentials")
        return self.access_token
    
    def _make_graph_request(self, url: str, stream: bool = False):
        """Make authenticated request to Microsoft Graph API"""
        if not self.access_token:
            self._get_access_token()
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'Accept': 'application/json'
        }
        
        response = requests.get(url, headers=headers, timeout=60, stream=stream)
        
        if response.status_code == 401:
            # Token might be expired, try to refresh
            self._get_access_token()
            headers['Authorization'] = f'Bearer {self.access_token}'
            response = requests.get(url, headers=headers, timeout=60, stream=stream)
        
        if response.status_code != 200:
            logger.error(f"Graph API request failed: {response.status_code} - {response.text}")
            raise Exception(f"Graph API request failed: {response.status_code}")
        
        if stream:
            return response.content
        return response.json()
    
    def _get_site_id(self) -> str:
        """Get the SharePoint site ID"""
        # Parse site URL to get hostname and site path
        # URL format: https://rgafarms.sharepoint.com/sites/Crops
        from urllib.parse import urlparse
        parsed = urlparse(self.site_url)
        hostname = parsed.netloc
        site_path = parsed.path
        
        # Get site by path
        url = f"{self.graph_url}/sites/{hostname}:{site_path}"
        site_info = self._make_graph_request(url)
        site_id = site_info['id']
        logger.info(f"Found site ID: {site_id}")
        return site_id
    
    def _get_drive_id(self, site_id: str) -> str:
        """Get the default document library drive ID for the site"""
        url = f"{self.graph_url}/sites/{site_id}/drives"
        drives = self._make_graph_request(url)
        
        if not drives.get('value'):
            raise Exception("No document libraries found in the site")
        
        # Use the first drive (usually "Documents")
        drive_id = drives['value'][0]['id']
        logger.info(f"Found drive ID: {drive_id}")
        return drive_id
    
    def _find_file(self, drive_id: str, filename: str) -> str:
        """Find a file in the drive by name, checking specific folder first"""
        
        # First try the specific folder path (Shared Documents/General)
        try:
            folder_url = f"{self.graph_url}/drives/{drive_id}/root:/{self.folder_path}:/children"
            items = self._make_graph_request(folder_url)
            
            for item in items.get('value', []):
                if item['name'].lower() == filename.lower():
                    logger.info(f"Found file in {self.folder_path}: {item['name']} (ID: {item['id']})")
                    return item['id']
        except Exception as e:
            logger.warning(f"Could not access folder {self.folder_path}: {e}")
        
        # Try root folder
        try:
            url = f"{self.graph_url}/drives/{drive_id}/root/children"
            items = self._make_graph_request(url)
            
            for item in items.get('value', []):
                if item['name'].lower() == filename.lower():
                    logger.info(f"Found file in root: {item['name']} (ID: {item['id']})")
                    return item['id']
        except Exception as e:
            logger.warning(f"Could not access root folder: {e}")
        
        # Search recursively as fallback
        try:
            url = f"{self.graph_url}/drives/{drive_id}/root/search(q='{filename}')"
            search_results = self._make_graph_request(url)
            
            for item in search_results.get('value', []):
                if item['name'].lower() == filename.lower():
                    logger.info(f"Found file via search: {item['name']} (ID: {item['id']})")
                    return item['id']
        except Exception as e:
            logger.warning(f"Search failed: {e}")
        
        raise Exception(f"File '{filename}' not found in SharePoint")
    
    def _download_file(self, drive_id: str, item_id: str) -> bytes:
        """Download file content from SharePoint"""
        url = f"{self.graph_url}/drives/{drive_id}/items/{item_id}/content"
        content = self._make_graph_request(url, stream=True)
        logger.info(f"Downloaded file: {len(content)} bytes")
        return content
    
    def _parse_staff_excel(self, file_content: bytes) -> List[Dict]:
        """Parse staff Excel file and extract employee data"""
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook[workbook.sheetnames[0]]
        
        # Get headers
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        logger.info(f"Excel headers: {headers}")
        
        # Find column indices - prioritize 'employee number' column
        name_col = None
        number_col = None
        workshop_col = None
        admin_col = None
        manager_col = None
        
        for i, header in enumerate(headers):
            # Check for employee number column FIRST (more specific match)
            if ('employee' in header and 'number' in header) or header == 'emp no' or header == 'employee_number':
                number_col = i
            elif 'name' in header and 'employee' not in header:
                name_col = i
            elif 'workshop' in header and 'control' in header:
                workshop_col = i
            elif 'admin' in header and 'control' in header:
                admin_col = i
            elif 'manager' in header:
                manager_col = i
        
        # If we didn't find employee number yet, look for other patterns (but NOT phone number)
        if number_col is None:
            for i, header in enumerate(headers):
                if ('number' in header or 'emp' in header) and 'phone' not in header and 'tel' not in header:
                    number_col = i
                    break
        
        # Fallback
        if name_col is None:
            name_col = 0
        if number_col is None and len(headers) > 1:
            number_col = 1
        
        logger.info(f"Column mapping - name: {name_col}, number: {number_col}, workshop: {workshop_col}, admin: {admin_col}, manager: {manager_col}")
        
        if number_col is None:
            raise Exception("Could not find Employee Number column")
        
        # Extract staff data
        staff_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > max(name_col, number_col):
                name = str(row[name_col]).strip() if row[name_col] else ''
                emp_number = str(row[number_col]).strip() if row[number_col] else ''
                
                workshop_control = None
                admin_control = None
                manager_control = None
                
                if workshop_col is not None and len(row) > workshop_col and row[workshop_col]:
                    workshop_control = str(row[workshop_col]).strip().lower()
                
                if admin_col is not None and len(row) > admin_col and row[admin_col]:
                    admin_control = str(row[admin_col]).strip().lower()
                
                if manager_col is not None and len(row) > manager_col and row[manager_col]:
                    manager_control = str(row[manager_col]).strip().lower()
                
                if name and emp_number and name.lower() not in ['name', 'staff', 'employee']:
                    staff_data.append({
                        'name': name,
                        'employee_number': emp_number,
                        'active': True,
                        'workshop_control': workshop_control,
                        'admin_control': admin_control,
                        'manager_control': manager_control
                    })
        
        logger.info(f"Parsed {len(staff_data)} staff members from Excel")
        return staff_data
    
    async def sync_staff_list(self, db) -> Dict:
        """Main sync function - downloads staff list from SharePoint and updates database"""
        try:
            logger.info(f"Starting SharePoint staff sync at {datetime.now()}")
            
            # Get site and drive info
            site_id = self._get_site_id()
            drive_id = self._get_drive_id(site_id)
            
            # Find and download the staff file
            item_id = self._find_file(drive_id, self.staff_filename)
            file_content = self._download_file(drive_id, item_id)
            
            # Parse the Excel file
            staff_data = self._parse_staff_excel(file_content)
            
            if not staff_data:
                raise Exception("No valid staff data found in Excel file")
            
            # Update database - preserve admin account (4444)
            from pydantic import BaseModel
            from typing import Optional
            
            class Staff(BaseModel):
                name: str
                employee_number: str
                active: bool = True
                workshop_control: Optional[str] = None
                admin_control: Optional[str] = None
                manager_control: Optional[str] = None
            
            delete_result = await db.staff.delete_many({"employee_number": {"$ne": "4444"}})
            logger.info(f"Deleted {delete_result.deleted_count} existing staff records")
            
            new_staff = [Staff(**data).dict() for data in staff_data]
            insert_result = await db.staff.insert_many(new_staff)
            logger.info(f"Inserted {len(insert_result.inserted_ids)} new staff records")
            
            result = {
                'success': True,
                'message': f'Successfully synced {len(staff_data)} staff members from SharePoint',
                'count': len(staff_data),
                'synced_at': datetime.now().isoformat(),
                'preview': staff_data[:5]
            }
            
            logger.info(f"SharePoint sync completed: {result['message']}")
            return result
            
        except Exception as e:
            logger.error(f"SharePoint sync failed: {str(e)}")
            return {
                'success': False,
                'message': f'Sync failed: {str(e)}',
                'synced_at': datetime.now().isoformat()
            }
    
    def _parse_assets_excel(self, file_content: bytes) -> Tuple[List[Dict], List[Dict]]:
        """Parse assets Excel file and extract asset data and checklist templates"""
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        
        # First sheet contains assets
        sheet = workbook[workbook.sheetnames[0]]
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        logger.info(f"Assets Excel headers: {headers}")
        
        # Find column indices
        check_type_col = None
        name_col = None
        make_col = None
        
        for i, header in enumerate(headers):
            if header == 'check type' or 'checktype' in header:
                check_type_col = i
            elif header == 'namecolumn' or ('name' in header and 'check' not in header):
                name_col = i
            elif header == 'makecolumn' or 'make' in header:
                make_col = i
        
        if check_type_col is None or name_col is None or make_col is None:
            raise Exception(f"Could not find required columns. Found: {headers}")
        
        logger.info(f"Assets column mapping - check_type: {check_type_col}, name: {name_col}, make: {make_col}")
        
        # Extract assets
        assets = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > max(check_type_col, name_col, make_col):
                check_type = str(row[check_type_col]).strip() if row[check_type_col] else ''
                name = str(row[name_col]).strip() if row[name_col] else ''
                make = str(row[make_col]).strip() if row[make_col] else ''
                
                if check_type and name and make:
                    assets.append({
                        'check_type': check_type,
                        'name': name,
                        'make': make
                    })
        
        logger.info(f"Parsed {len(assets)} assets from Excel")
        
        # Process checklist template sheets
        checklist_templates = []
        unique_check_types = set(asset['check_type'] for asset in assets)
        
        for sheet_name in workbook.sheetnames[1:]:  # Skip first sheet (assets)
            sheet = workbook[sheet_name]
            
            # Try to match sheet name with check types
            matching_check_type = None
            sheet_name_clean = sheet_name.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '').replace('checklist', '')
            
            for check_type in unique_check_types:
                check_type_clean = check_type.lower().replace('/', '').replace(' ', '').replace('_', '').replace('-', '').replace('checklist', '')
                if sheet_name_clean == check_type_clean or check_type.lower() == sheet_name.lower():
                    matching_check_type = check_type
                    break
            
            # Fuzzy matching fallback
            if not matching_check_type:
                for check_type in unique_check_types:
                    if check_type_clean in sheet_name_clean or sheet_name_clean in check_type_clean:
                        matching_check_type = check_type
                        break
            
            if not matching_check_type:
                logger.warning(f"Sheet '{sheet_name}' doesn't match any check type, skipping")
                continue
            
            # Extract checklist items from sheet
            items = []
            sheet_headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
            
            item_col = None
            critical_col = None
            photo_col = None
            
            for i, h in enumerate(sheet_headers):
                if 'item' in h or 'check' in h or 'task' in h:
                    item_col = i
                elif 'critical' in h or 'common' in h:
                    critical_col = i
                elif 'photo' in h:
                    photo_col = i
            
            if item_col is None:
                item_col = 0  # Fallback to first column
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > item_col and row[item_col]:
                    item_text = str(row[item_col]).strip()
                    if item_text and item_text.lower() not in ['item', 'check', 'task', '']:
                        is_critical = False
                        photo_required = False
                        
                        if critical_col is not None and len(row) > critical_col and row[critical_col]:
                            is_critical = str(row[critical_col]).strip().lower() in ['yes', 'true', '1', 'y']
                        
                        if photo_col is not None and len(row) > photo_col and row[photo_col]:
                            photo_required = str(row[photo_col]).strip().lower() in ['yes', 'true', '1', 'y']
                        
                        items.append({
                            'item': item_text,
                            'critical': is_critical,
                            'photo_required': photo_required
                        })
            
            if items:
                checklist_templates.append({
                    'check_type': matching_check_type,
                    'sheet_name': sheet_name,
                    'items': items
                })
                logger.info(f"Parsed {len(items)} checklist items for '{matching_check_type}' from sheet '{sheet_name}'")
        
        return assets, checklist_templates
    
    async def sync_assets_list(self, db) -> Dict:
        """Sync assets and checklist templates from SharePoint"""
        try:
            logger.info(f"Starting SharePoint assets sync at {datetime.now()}")
            
            # Get site and drive info
            site_id = self._get_site_id()
            drive_id = self._get_drive_id(site_id)
            
            # Find and download the assets file
            item_id = self._find_file(drive_id, self.assets_filename)
            file_content = self._download_file(drive_id, item_id)
            
            # Parse the Excel file
            assets, checklist_templates = self._parse_assets_excel(file_content)
            
            if not assets:
                raise Exception("No valid asset data found in Excel file")
            
            # Get existing QR print status to preserve
            existing_assets = await db.assets.find({}, {"_id": 0}).to_list(length=10000)
            existing_qr_status = {}
            for ea in existing_assets:
                key = f"{ea.get('make', '')}:{ea.get('name', '')}"
                if ea.get('qr_printed'):
                    existing_qr_status[key] = {
                        'qr_printed': ea.get('qr_printed', False),
                        'qr_printed_at': ea.get('qr_printed_at')
                    }
            
            # Update assets database
            await db.assets.delete_many({})
            
            import uuid
            new_assets = []
            for asset in assets:
                asset_dict = {
                    'id': str(uuid.uuid4()),
                    'check_type': asset['check_type'],
                    'name': asset['name'],
                    'make': asset['make'],
                    'qr_printed': False,
                    'qr_printed_at': None
                }
                # Preserve QR print status
                key = f"{asset_dict['make']}:{asset_dict['name']}"
                if key in existing_qr_status:
                    asset_dict['qr_printed'] = existing_qr_status[key]['qr_printed']
                    asset_dict['qr_printed_at'] = existing_qr_status[key]['qr_printed_at']
                new_assets.append(asset_dict)
            
            await db.assets.insert_many(new_assets)
            logger.info(f"Inserted {len(new_assets)} assets")
            
            # Update checklist templates
            templates_count = 0
            for template in checklist_templates:
                await db.checklist_templates.update_one(
                    {'check_type': template['check_type']},
                    {'$set': {
                        'check_type': template['check_type'],
                        'sheet_name': template['sheet_name'],
                        'items': template['items'],
                        'updated_at': datetime.now().isoformat()
                    }},
                    upsert=True
                )
                templates_count += 1
            
            logger.info(f"Updated {templates_count} checklist templates")
            
            result = {
                'success': True,
                'message': f'Successfully synced {len(assets)} assets and {templates_count} checklist templates',
                'assets_count': len(assets),
                'templates_count': templates_count,
                'synced_at': datetime.now().isoformat(),
                'preview': assets[:5]
            }
            
            logger.info(f"Assets sync completed: {result['message']}")
            return result
            
        except Exception as e:
            logger.error(f"Assets sync failed: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return {
                'success': False,
                'message': f'Sync failed: {str(e)}',
                'synced_at': datetime.now().isoformat()
            }
    
    async def sync_all(self, db) -> Dict:
        """Sync both staff and assets from SharePoint"""
        staff_result = await self.sync_staff_list(db)
        assets_result = await self.sync_assets_list(db)
        
        return {
            'success': staff_result.get('success', False) and assets_result.get('success', False),
            'staff': staff_result,
            'assets': assets_result,
            'synced_at': datetime.now().isoformat()
        }
    
    def test_connection(self) -> Dict:
        """Test the SharePoint connection for both files"""
        # Always include credentials info for debugging
        credentials_info = {
            'client_id_prefix': self.client_id[:8] if self.client_id else 'NONE',
            'tenant_id_prefix': self.tenant_id[:8] if self.tenant_id else 'NONE',
            'secret_length': len(self.client_secret) if self.client_secret else 0,
            'secret_prefix': self.client_secret[:4] if self.client_secret else 'NONE'
        }
        
        try:
            # Log credential info (not the actual values)
            logger.info(f"Testing connection with client_id starting with: {self.client_id[:8] if self.client_id else 'NONE'}...")
            logger.info(f"Tenant ID starting with: {self.tenant_id[:8] if self.tenant_id else 'NONE'}...")
            logger.info(f"Secret length: {len(self.client_secret) if self.client_secret else 0}")
            
            self._get_access_token()
            site_id = self._get_site_id()
            drive_id = self._get_drive_id(site_id)
            
            result = {
                'success': True,
                'message': 'SharePoint connection successful',
                'site_id': site_id,
                'drive_id': drive_id,
                'files': {},
                'credentials_info': credentials_info
            }
            
            # Check staff file
            try:
                staff_item_id = self._find_file(drive_id, self.staff_filename)
                url = f"{self.graph_url}/drives/{drive_id}/items/{staff_item_id}"
                staff_info = self._make_graph_request(url)
                result['files']['staff'] = {
                    'file_name': staff_info.get('name'),
                    'file_size': staff_info.get('size'),
                    'last_modified': staff_info.get('lastModifiedDateTime'),
                    'status': 'found'
                }
            except Exception as e:
                result['files']['staff'] = {'status': 'not_found', 'error': str(e)}
            
            # Check assets file
            try:
                assets_item_id = self._find_file(drive_id, self.assets_filename)
                url = f"{self.graph_url}/drives/{drive_id}/items/{assets_item_id}"
                assets_info = self._make_graph_request(url)
                result['files']['assets'] = {
                    'file_name': assets_info.get('name'),
                    'file_size': assets_info.get('size'),
                    'last_modified': assets_info.get('lastModifiedDateTime'),
                    'status': 'found'
                }
            except Exception as e:
                result['files']['assets'] = {'status': 'not_found', 'error': str(e)}
            
            return result
        except Exception as e:
            return {
                'success': False,
                'message': f'Connection failed: {str(e)}',
                'credentials_info': credentials_info
            }


# Global instance
sharepoint_auto_sync = SharePointAutoSync()
